import csv
import json
import logging
import math
import os
import re
import sys
import uuid
import sqlite3
import threading
import time
import unicodedata
from collections import Counter
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from functools import wraps
from io import BytesIO, StringIO
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import requests
from dotenv import load_dotenv
from flask import Flask, abort, flash, has_request_context, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from flask_migrate import Migrate
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, case, or_
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.datastructures import FileStorage
import click

try:
    from .models import EquipaNota, GeocodingCache, HistoricoLead, Lead, PlanoReunioes, PossivelDuplicado, User, db
except ImportError:
    from models import EquipaNota, GeocodingCache, HistoricoLead, Lead, PlanoReunioes, PossivelDuplicado, User, db


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
DB_PATH = os.path.join(INSTANCE_DIR, "decisao_comercial.db")
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_USER_AGENT = "AllteraLeadPlanner/1.0 academico"
# Em ambiente académico/local é preferível não bloquear o fluxo de importação
# com um lock global entre processos do reloader Flask. O POST continua a usar
# transação/rollback, mas uma importação interrompida já não deixa a UI presa.
IMPORT_LOCK = None
IMPORT_JOBS = {}
IMPORT_JOBS_LOCK = threading.Lock()
AUDIT_CONTEXT = threading.local()

load_dotenv()
migrate = Migrate()
login_manager = LoginManager()
logger = logging.getLogger(__name__)
IMPORT_UPLOADS = {}
IMPORT_UPLOADS_LOCK = threading.Lock()

LEAD_STATES = [
    "Por contactar",
    "Ligar de volta",
    "Adiar contacto",
    "Já tratado / no CRM",
    "Sem interesse",
]
ACTIVE_STATES = ["Por contactar", "Ligar de volta"]
SCHEDULED_STATES = {"Ligar de volta", "Adiar contacto", "Sem interesse"}
MAP_INACTIVE_STATE_KEYS = {
    "inativa",
    "inativo",
    "naoativa",
    "naoativo",
    "desativada",
    "desativado",
    "arquivada",
    "arquivado",
    "perdida",
    "perdido",
    "seminteresse",
    "naointeressada",
    "naointeressado",
    "tratadonocrm",
    "tratadocrm",
    "duplicada",
    "duplicado",
}
CLASSIFICATION_FILTERS = [
    ("ativos", "Todos ativos"),
    ("por_contactar", "Por contactar"),
    ("ligar_volta", "Ligar de volta"),
    ("adiados", "Adiados"),
    ("com_observacoes", "Com observações"),
    ("crm", "Já tratado / no CRM"),
    ("sem_interesse", "Sem interesse"),
]
COMMERCIALS = ["FlÃ¡via", "Bernardo", "Outro"]
UNASSIGNED_COMMERCIAL = "Sem comercial atribuído"
PLANNING_COMMERCIALS = ["Inês", "Bruno", "Flávia", "Miriam", "Setil"]
MAP_COMMERCIALS = [
    {"value": "", "label": "Todos"},
    {"value": "sem_comercial", "label": "Sem comercial"},
    {"value": "ines", "label": "Inês"},
    {"value": "bruno", "label": "Bruno"},
    {"value": "flavia", "label": "Flávia"},
    {"value": "miriam", "label": "Miriam"},
    {"value": "setil", "label": "Setil"},
]
TAG_OPTIONS = [
    "Quente",
    "Frio",
    "Urgente",
    "VIP",
    "Recuperar",
    "Sem comercial",
    "Restaurante",
    "Hotel",
    "Cadeia",
    "Grande dimensÃ£o",
    "Pequeno negÃ³cio",
    "Contacto difÃ­cil",
    "Voltar mais tarde",
    "Prioridade operacional",
]
INSIGHT_TAG_OPTIONS = [
    "VIP",
    "Difícil",
    "Impaciente",
    "Só WhatsApp",
    "Prefere email",
    "Follow-up delicado",
    "Interesse elevado",
    "Sem resposta",
    "Gatekeeper",
    "Recuperar",
]
INVALID_CITY_VALUES = {"", "-", "n/a", "na", "null", "none", "sem cidade", "desconhecido"}
CITY_CORRECTIONS = {
    "lisboas": "Lisboa",
    "setuball": "Setúbal",
    "albufeira algarve": "Albufeira",
    "lagoa algarve": "Lagoa",
}
REGION_WORDS = {"algarve", "alentejo", "portugal"}
# Regiões genéricas em segmentos compostos (ex.: "Açores - Ponta Delgada").
# Nota: "algarve" não está aqui — células só com Algarve usam REGION_FALLBACK_COORDS.
GEO_GENERIC_REGIONS = {
    "alentejo",
    "norte",
    "centro",
    "sul",
    "portugal",
    "regiao",
    "zona",
    "acores",
    "madeira",
    "ilhas",
}
LOCALITY_FALLBACK_SEQUENCES = {}
FIELD_ALIASES = {
    "nome_empresa": ["nome_empresa", "empresa", "nome", "cliente", "nome_cliente"],
    "telefone": ["telefone", "telemovel", "telemÃ³vel", "contacto_telefonico", "contacto_telefÃ³nico"],
    "email": ["email", "e-mail", "mail"],
    "morada": ["morada", "endereco", "endereÃ§o", "rua"],
    "localidade": ["cidade", "localidade", "concelho"],
    "tipo_cliente": ["tipo", "categoria", "segmento", "tipo_cliente"],
    "codigo_postal": ["codigo_postal", "cÃ³digo_postal", "cp"],
    "contacto": ["contacto", "pessoa_contacto", "responsavel"],
    "observacoes": ["observacoes", "observaÃ§Ãµes", "notas"],
    "latitude": ["latitude", "lat"],
    "longitude": ["longitude", "lng", "lon"],
}
IMPORT_COLUMNS = [
    "nome_empresa",
    "tipo_cliente",
    "morada",
    "codigo_postal",
    "localidade",
    "contacto",
    "telefone",
    "email",
    "observacoes",
    "latitude",
    "longitude",
]
# Campos extra usados nos ficheiros reais da Alltera. Mantemos o formato antigo
# e acrescentamos estes campos ao mapeamento confirmado pelo utilizador.
FIELD_ALIASES.update({
    "nome_empresa": ["nome_empresa", "empresa", "nome", "cliente", "nome cliente", "nome_cliente", "designacao", "designação"],
    "telefone": ["telefone", "telemovel", "telemóvel", "contacto_telefonico", "contacto telefonico", "contacto telefónico", "contacto", "tel", "tlm", "mobile"],
    "tipo_cliente": ["area_de_negocio", "area de negocio", "área_de_negócio", "área de negócio", "tipo", "segmento", "tipo_cliente", "assist. t.", "assist t", "assistencia", "assistência"],
    "categoria": ["categoria", "assist. t.", "assist t", "assistencia", "assistência"],
    "morada": ["morada", "endereco", "endereço", "rua", "morada completa", "morada_completa"],
    "codigo_postal": ["codigo_postal", "código_postal", "codigo postal", "código postal", "cod. p.", "cod p", "cod postal", "cp", "c.p."],
    "localidade": ["cidade", "localidade", "concelho", "zona", "regiao", "região"],
    "contacto": ["nome_cliente", "nome cliente", "contacto", "pessoa_contacto", "responsavel", "responsável"],
    "comercial_responsavel": ["comercial", "comercial_responsavel"],
    "nif": ["nif", "numero_fiscal", "numero fiscal", "número fiscal"],
    "observacoes": ["observacoes", "observações", "obs", "notas", "nota"],
    "observacoes_contacto": ["observacoes_do_contacto", "observacoes do contacto", "observações_do_contacto", "observações do contacto", "notas_contacto"],
    "reuniao_info": ["reuniao", "reunião"],
    "dia": ["dia", "day"],
    "mes": ["mes", "mês", "month"],
    "ano": ["ano", "year"],
})
IMPORT_COLUMNS.extend([
    "comercial_responsavel",
    "categoria",
    "nif",
    "observacoes_contacto",
    "reuniao_info",
    "dia",
    "mes",
    "ano",
])
ALLTERA_REAL_COLUMNS = [
    "Comercial",
    "Categoria",
    "Ãrea de negÃ³cio",
    "Nome Cliente",
    "Contacto telefÃ³nico",
    "email",
    "Empresa",
    "NIF",
    "Cidade",
    "ObservaÃ§Ãµes",
    "ReuniÃ£o",
    "ObservaÃ§Ãµes do contacto",
]
ALLTERA_REAL_COLUMNS = [
    "Comercial",
    "Categoria",
    "\u00c1rea de neg\u00f3cio",
    "Nome Cliente",
    "Contacto telef\u00f3nico",
    "email",
    "Empresa",
    "NIF",
    "Cidade",
    "Observa\u00e7\u00f5es",
    "Reuni\u00e3o",
    "Observa\u00e7\u00f5es do contacto",
]
NOMINATIM_USER_AGENT = "AllteraLeadPlanner/1.0 acad\u00e9mico"
REQUIRED_COLUMNS = []


def get_database_uri():
    database_url = os.getenv("DATABASE_URL")
    if database_url:
        database_url = database_url.replace("postgres://", "postgresql://", 1)
        parsed_url = urlsplit(database_url)
        query_params = dict(parse_qsl(parsed_url.query, keep_blank_values=True))
        hostname = parsed_url.hostname or ""
        if "supabase.com" in hostname.lower() and "sslmode" not in query_params:
            query_params["sslmode"] = "require"
            database_url = urlunsplit(parsed_url._replace(query=urlencode(query_params)))
        return database_url
    return f"sqlite:///{DB_PATH}"


def is_safe_next_url(target):
    if not target:
        return False
    parsed = urlsplit(target)
    return not parsed.netloc and parsed.path.startswith("/") and not parsed.path.startswith("//")


def is_flask_db_command():
    return "db" in sys.argv[1:]


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped_view(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                abort(403)
            return view(*args, **kwargs)

        return wrapped_view

    return decorator


admin_required = role_required("admin")


def active_commercial_users():
    return User.query.filter_by(ativo=True, role="comercial").order_by(User.nome.asc()).all()


def assignment_scope_options():
    return [
        {"value": "all", "label": "Todas"},
        {"value": "por_contactar", "label": "Por contactar"},
        {"value": "contactadas", "label": "Contactadas"},
        {"value": "agendadas", "label": "Agendadas"},
        {"value": "sem_proximo_contacto", "label": "Sem próximo contacto"},
    ]


def requested_assignment_scope(default=None):
    if default is None:
        default = "all"
    scope = request.args.get("scope", default)
    allowed = {option["value"] for option in assignment_scope_options()}
    return scope if scope in allowed else default


def apply_assignment_scope(query, scope):
    if scope == "por_contactar":
        return query.filter(Lead.estado == "Por contactar")
    if scope == "contactadas":
        return query.filter(Lead.estado != "Por contactar")
    if scope == "agendadas":
        return query.filter(Lead.data_novo_contacto.isnot(None))
    if scope == "sem_proximo_contacto":
        return query.filter(Lead.data_novo_contacto.is_(None))
    return query


def create_user_record(nome, email, password, role):
    normalized_role = clean_text(role).lower() or "comercial"
    if normalized_role not in {"admin", "comercial"}:
        raise click.ClickException("Role invalida. Usa 'admin' ou 'comercial'.")

    normalized_email = clean_text(email).lower()
    existing = User.query.filter_by(email=normalized_email).first()
    if existing:
        raise click.ClickException("Ja existe um utilizador com esse email.")

    user = User(nome=clean_text(nome), email=normalized_email, role=normalized_role, ativo=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return user


def create_app():
    os.makedirs(INSTANCE_DIR, exist_ok=True)
    app = Flask(__name__, instance_path=INSTANCE_DIR, instance_relative_config=True)
    app.config["SECRET_KEY"] = "mudar_esta_chave"
    app.config["SQLALCHEMY_DATABASE_URI"] = get_database_uri()
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}
    if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite:"):
        app.config["SQLALCHEMY_ENGINE_OPTIONS"]["connect_args"] = {
            "timeout": 30,
            "check_same_thread": False,
        }

    db.init_app(app)
    migrate.init_app(app, db)
    login_manager.init_app(app)
    login_manager.login_view = "login"
    login_manager.login_message = "Inicia sessao para aceder a Alltera."
    login_manager.login_message_category = "warning"
    with app.app_context():
        if not is_flask_db_command():
            if db.engine.dialect.name == "sqlite":
                db.create_all()
                configure_sqlite_connection()
            migrate_database()

    @login_manager.user_loader
    def load_user(user_id):
        try:
            return db.session.get(User, int(user_id))
        except (TypeError, ValueError):
            return None

    @app.cli.command("create-admin")
    @click.option("--nome", required=True, help="Nome do utilizador.")
    @click.option("--email", required=True, help="Email de login.")
    @click.option("--password", required=True, help="Password inicial.")
    def create_admin(nome, email, password):
        user = create_user_record(nome=nome, email=email, password=password, role="admin")
        click.echo(f"Admin criado: {user.email}")

    @app.cli.command("create-user")
    @click.option("--nome", required=True, help="Nome do utilizador.")
    @click.option("--email", required=True, help="Email de login.")
    @click.option("--password", required=True, help="Password inicial.")
    @click.option("--role", default="comercial", show_default=True, help="Role: admin ou comercial.")
    def create_user(nome, email, password, role):
        user = create_user_record(nome=nome, email=email, password=password, role=role)
        click.echo(f"Utilizador criado: {user.email} ({user.role})")

    register_routes(app)
    return app


def configure_sqlite_connection():
    db.session.execute(db.text("PRAGMA busy_timeout = 30000"))
    db.session.execute(db.text("PRAGMA journal_mode = WAL"))
    db.session.commit()


def rebuild_legacy_lead_table_if_needed():
    schema = db.session.execute(db.text("PRAGMA table_info(lead)")).fetchall()
    if not schema:
        return
    columns = {row[1]: row for row in schema}
    legacy_required = any(columns.get(name) and columns[name][3] for name in ["latitude", "longitude", "prioridade", "estado_lead"])
    if not legacy_required:
        return

    target_columns = [
        ("id", "INTEGER PRIMARY KEY"),
        ("nome_cliente", "VARCHAR(180)"),
        ("area_negocio", "VARCHAR(120)"),
        ("cidade", "VARCHAR(80)"),
        ("empresa", "VARCHAR(180)"),
        ("nome_empresa", "VARCHAR(180) NOT NULL"),
        ("tipo_cliente", "VARCHAR(80) NOT NULL"),
        ("morada", "VARCHAR(220)"),
        ("codigo_postal", "VARCHAR(20)"),
        ("localidade", "VARCHAR(80) NOT NULL"),
        ("contacto", "VARCHAR(120)"),
        ("telefone", "VARCHAR(40)"),
        ("email", "VARCHAR(160)"),
        ("categoria", "VARCHAR(120)"),
        ("nif", "VARCHAR(30)"),
        ("observacoes", "TEXT"),
        ("observacoes_contacto", "TEXT"),
        ("reuniao_info", "TEXT"),
        ("classificacao_observacao", "VARCHAR(80)"),
        ("motivo_classificacao", "TEXT"),
        ("latitude", "FLOAT"),
        ("longitude", "FLOAT"),
        ("estado", "VARCHAR(40) NOT NULL DEFAULT 'Por contactar'"),
        ("estado_lead", "VARCHAR(40) DEFAULT 'Por contactar'"),
        ("prioridade", "VARCHAR(20) DEFAULT 'Baixa'"),
        ("comercial_responsavel", "VARCHAR(80) NOT NULL DEFAULT 'Outro'"),
        ("data_novo_contacto", "DATE"),
        ("data_reuniao", "DATE"),
        ("hora_reuniao", "VARCHAR(10)"),
        ("tags", "VARCHAR(300)"),
        ("insight_tags", "VARCHAR(400)"),
        ("insight_note", "TEXT"),
        ("created_at", "DATETIME"),
        ("updated_at", "DATETIME"),
    ]
    db.session.execute(db.text(f"CREATE TABLE lead_new ({', '.join(f'{name} {definition}' for name, definition in target_columns)})"))
    select_parts = []
    for name, _definition in target_columns:
        if name in columns:
            select_parts.append(name)
        elif name == "estado":
            select_parts.append("'Por contactar'")
        elif name == "estado_lead":
            select_parts.append("'Por contactar'")
        elif name == "prioridade":
            select_parts.append("'Baixa'")
        elif name == "comercial_responsavel":
            select_parts.append("'Outro'")
        else:
            select_parts.append("NULL")
    column_names = ", ".join(name for name, _definition in target_columns)
    db.session.execute(db.text(f"INSERT INTO lead_new ({column_names}) SELECT {', '.join(select_parts)} FROM lead"))
    db.session.execute(db.text("DROP TABLE lead"))
    db.session.execute(db.text("ALTER TABLE lead_new RENAME TO lead"))
    db.session.commit()


def migrate_database():
    dialect = db.engine.dialect.name
    if dialect == "sqlite":
        user_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(users)")).fetchall()}
        user_migrations = {
            "approval_status": "ALTER TABLE users ADD COLUMN approval_status TEXT NOT NULL DEFAULT 'approved'",
            "approved_at": "ALTER TABLE users ADD COLUMN approved_at DATETIME",
            "approved_by_id": "ALTER TABLE users ADD COLUMN approved_by_id INTEGER",
        }
    elif dialect == "postgresql":
        user_columns = {
            row[0]
            for row in db.session.execute(db.text(
                "SELECT column_name FROM information_schema.columns WHERE table_name='users'"
            )).fetchall()
        }
        user_migrations = {
            "approval_status": "ALTER TABLE users ADD COLUMN approval_status VARCHAR(20) NOT NULL DEFAULT 'approved'",
            "approved_at": "ALTER TABLE users ADD COLUMN approved_at TIMESTAMP NULL",
            "approved_by_id": "ALTER TABLE users ADD COLUMN approved_by_id INTEGER NULL",
        }
    else:
        user_columns = set()
        user_migrations = {}

    for column, sql in user_migrations.items():
        if column not in user_columns:
            db.session.execute(db.text(sql))
    db.session.commit()

    if dialect == "sqlite":
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS equipa_nota (
                id INTEGER PRIMARY KEY,
                titulo VARCHAR(140) NOT NULL,
                conteudo TEXT NOT NULL,
                autor_id INTEGER,
                created_at DATETIME NOT NULL,
                updated_at DATETIME NOT NULL,
                FOREIGN KEY(autor_id) REFERENCES users (id)
            )
        """))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_equipa_nota_autor_id ON equipa_nota (autor_id)"))
        db.session.commit()
    elif dialect == "postgresql":
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS equipa_nota (
                id SERIAL PRIMARY KEY,
                titulo VARCHAR(140) NOT NULL,
                conteudo TEXT NOT NULL,
                autor_id INTEGER REFERENCES users (id),
                created_at TIMESTAMP NOT NULL,
                updated_at TIMESTAMP NOT NULL
            )
        """))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_equipa_nota_autor_id ON equipa_nota (autor_id)"))
        db.session.commit()

    if dialect != "sqlite":
        return

    rebuild_legacy_lead_table_if_needed()
    columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(lead)")).fetchall()}
    migrations = {
        "nome_cliente": "ALTER TABLE lead ADD COLUMN nome_cliente VARCHAR(180)",
        "area_negocio": "ALTER TABLE lead ADD COLUMN area_negocio VARCHAR(120)",
        "cidade": "ALTER TABLE lead ADD COLUMN cidade VARCHAR(80)",
        "empresa": "ALTER TABLE lead ADD COLUMN empresa VARCHAR(180)",
        "contacto": "ALTER TABLE lead ADD COLUMN contacto VARCHAR(120)",
        "telefone": "ALTER TABLE lead ADD COLUMN telefone VARCHAR(40)",
        "email": "ALTER TABLE lead ADD COLUMN email VARCHAR(160)",
        "categoria": "ALTER TABLE lead ADD COLUMN categoria VARCHAR(120)",
        "nif": "ALTER TABLE lead ADD COLUMN nif VARCHAR(30)",
        "observacoes": "ALTER TABLE lead ADD COLUMN observacoes TEXT",
        "observacoes_contacto": "ALTER TABLE lead ADD COLUMN observacoes_contacto TEXT",
        "reuniao_info": "ALTER TABLE lead ADD COLUMN reuniao_info TEXT",
        "classificacao_observacao": "ALTER TABLE lead ADD COLUMN classificacao_observacao VARCHAR(80)",
        "motivo_classificacao": "ALTER TABLE lead ADD COLUMN motivo_classificacao TEXT",
        "latitude": "ALTER TABLE lead ADD COLUMN latitude FLOAT",
        "longitude": "ALTER TABLE lead ADD COLUMN longitude FLOAT",
        "estado": "ALTER TABLE lead ADD COLUMN estado VARCHAR(40) DEFAULT 'Por contactar'",
        "estado_lead": "ALTER TABLE lead ADD COLUMN estado_lead VARCHAR(40) DEFAULT 'Por contactar'",
        "prioridade": "ALTER TABLE lead ADD COLUMN prioridade VARCHAR(20) DEFAULT 'Baixa'",
        "comercial_responsavel": "ALTER TABLE lead ADD COLUMN comercial_responsavel VARCHAR(80) DEFAULT 'FlÃ¡via'",
        "data_novo_contacto": "ALTER TABLE lead ADD COLUMN data_novo_contacto DATE",
        "data_reuniao": "ALTER TABLE lead ADD COLUMN data_reuniao DATE",
        "hora_reuniao": "ALTER TABLE lead ADD COLUMN hora_reuniao VARCHAR(10)",
        "updated_at": "ALTER TABLE lead ADD COLUMN updated_at DATETIME",
        "tags": "ALTER TABLE lead ADD COLUMN tags VARCHAR(300)",
        "insight_tags": "ALTER TABLE lead ADD COLUMN insight_tags VARCHAR(400)",
        "insight_note": "ALTER TABLE lead ADD COLUMN insight_note TEXT",
        "assigned_to_id": "ALTER TABLE lead ADD COLUMN assigned_to_id INTEGER",
    }

    for column, sql in migrations.items():
        if column not in columns:
            db.session.execute(db.text(sql))

    legacy_map = {
        "name": "nome_empresa",
        "city": "localidade",
        "sector": "tipo_cliente",
        "status": "estado",
    }
    refreshed = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(lead)")).fetchall()}
    for old, new in legacy_map.items():
        if old in refreshed and new in refreshed:
            db.session.execute(db.text(f"UPDATE lead SET {new} = COALESCE({new}, {old})"))

    db.session.execute(db.text("UPDATE lead SET estado = 'Por contactar' WHERE estado IS NULL OR estado = '' OR estado IN ('Ativo', 'Prospect')"))
    db.session.execute(db.text("UPDATE lead SET estado_lead = COALESCE(estado_lead, estado, 'Por contactar')"))
    db.session.execute(db.text("UPDATE lead SET prioridade = 'Baixa' WHERE prioridade IS NULL OR prioridade = ''"))
    db.session.execute(db.text("UPDATE lead SET comercial_responsavel = :unassigned WHERE comercial_responsavel IS NULL OR comercial_responsavel = ''"), {"unassigned": UNASSIGNED_COMMERCIAL})
    db.session.execute(db.text("UPDATE lead SET nome_cliente = COALESCE(nome_cliente, contacto, nome_empresa)"))
    db.session.execute(db.text("UPDATE lead SET area_negocio = COALESCE(area_negocio, tipo_cliente, categoria, 'Outro')"))
    db.session.execute(db.text("UPDATE lead SET cidade = COALESCE(cidade, localidade, 'Sem cidade')"))
    db.session.execute(db.text("UPDATE lead SET empresa = COALESCE(empresa, nome_empresa)"))
    db.session.commit()

    history_tables = {row[0] for row in db.session.execute(db.text("SELECT name FROM sqlite_master WHERE type='table'")).fetchall()}
    if "historico_lead" in history_tables:
        history_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(historico_lead)")).fetchall()}
        history_migrations = {
            "user_id": "ALTER TABLE historico_lead ADD COLUMN user_id INTEGER",
            "tipo_acao": "ALTER TABLE historico_lead ADD COLUMN tipo_acao VARCHAR(80)",
            "resultado": "ALTER TABLE historico_lead ADD COLUMN resultado VARCHAR(120)",
        }
        for column, sql in history_migrations.items():
            if column not in history_columns:
                db.session.execute(db.text(sql))
        db.session.commit()

    for lead in Lead.query.all():
        lead.estado = normalize_legacy_state(lead.estado)
        lead.estado_lead = normalize_legacy_state(lead.estado_lead or lead.estado)
        if lead.estado not in LEAD_STATES:
            lead.estado = "Por contactar"
        if not lead.comercial_responsavel:
            lead.comercial_responsavel = UNASSIGNED_COMMERCIAL
        if lead.tipo_cliente:
            lead.tipo_cliente = normalize_client_type(lead.tipo_cliente)
        if lead.localidade:
            lead.localidade = normalize_locality(lead.localidade)
    db.session.commit()


def normalize_key(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.replace(" ", "_")


def compact_key(value):
    return re.sub(r"[^a-z0-9]", "", normalize_key(value))


def normalize_lookup(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


LOCALITY_FALLBACK_SEQUENCES.update({
    normalize_lookup("Borba - Alentejo"): ["Borba"],
    normalize_lookup("Ferrel - Peniche"): ["Ferrel", "Peniche"],
})


GEO_CITY_MANUAL = {
    normalize_lookup("Troia"): "Troia",
    normalize_lookup("Tróia"): "Troia",
    normalize_lookup("Estoril"): "Cascais",
    normalize_lookup("Alcabideche"): "Cascais",
    normalize_lookup("Algueirão"): "Sintra",
    normalize_lookup("Algueirao"): "Sintra",
    normalize_lookup("Alfragide"): "Amadora",
    normalize_lookup("Linda a Velha"): "Oeiras",
    normalize_lookup("Algés"): "Oeiras",
    normalize_lookup("Alges"): "Oeiras",
    normalize_lookup("Alg?s"): "Oeiras",
    normalize_lookup("Rio Mouro"): "Sintra",
    normalize_lookup("Serra Luz"): "Odivelas",
    normalize_lookup("Bom Sucesso"): "Lisboa",
    normalize_lookup("Alfornelos"): "Amadora",
    normalize_lookup("Carcavelos"): "Cascais",
    normalize_lookup("Paço de Arcos"): "Oeiras",
    normalize_lookup("Paco de Arcos"): "Oeiras",
    normalize_lookup("Carnaxide"): "Oeiras",
    normalize_lookup("Queluz"): "Sintra",
    normalize_lookup("Cacém"): "Sintra",
    normalize_lookup("Cacem"): "Sintra",
    normalize_lookup("Charneca da Caparica"): "Almada",
    normalize_lookup("Costa da Caparica"): "Almada",
    normalize_lookup("Quinta do Conde"): "Sesimbra",
    normalize_lookup("Olhos de Água"): "Albufeira",
    normalize_lookup("Olhos de Agua"): "Albufeira",
    normalize_lookup("Algarve"): "Faro",
    normalize_lookup("Ilha da Berlengas"): "Peniche",
    normalize_lookup("Set?bal"): "Setúbal",
    normalize_lookup("Setubal"): "Setúbal",
    normalize_lookup("Santarem"): "Santarém",
    normalize_lookup("Santar?m"): "Santarém",
    normalize_lookup("Vale Milhaços"): "Corroios",
    normalize_lookup("Vale Milhacos"): "Corroios",
    normalize_lookup("Leça do Balio"): "Matosinhos",
    normalize_lookup("Leca do Balio"): "Matosinhos",
    normalize_lookup("Ferreiras"): "Albufeira",
}


def _build_region_fallback_coords():
    """Coordenadas fixas por região genérica (antes do Nominatim). Chaves = normalize_lookup do texto."""
    out = {}
    entries = [
        ("Algarve", "Algarve", 37.0194, -7.9304),
        ("Região do Algarve", "Algarve", 37.0194, -7.9304),
        ("Regiao do Algarve", "Algarve", 37.0194, -7.9304),
        ("Zona Algarve", "Algarve", 37.0194, -7.9304),
        ("Algarve, Portugal", "Algarve", 37.0194, -7.9304),
        ("Algarve Portugal", "Algarve", 37.0194, -7.9304),
        ("Faro", "Faro", 37.0194, -7.9304),
        ("Oeste", "Oeste", 39.3600, -9.1500),
        ("Santarem", "Santarém", 39.2369, -8.6850),
        ("Santarém", "Santarém", 39.2369, -8.6850),
        ("Lisboa", "Lisboa", 38.7223, -9.1393),
        ("Amadora", "Amadora", 38.7597, -9.2397),
        ("Oeiras", "Oeiras", 38.6979, -9.3015),
        ("Cascais", "Cascais", 38.6968, -9.4215),
        ("Sintra", "Sintra", 38.8029, -9.3817),
        ("Odivelas", "Odivelas", 38.7927, -9.1838),
        ("Almada", "Almada", 38.6765, -9.1651),
        ("Sesimbra", "Sesimbra", 38.4445, -9.1015),
        ("Setubal", "Setúbal", 38.5244, -8.8882),
        ("Setúbal", "Setúbal", 38.5244, -8.8882),
        ("Albufeira", "Albufeira", 37.0891, -8.2479),
        ("Peniche", "Peniche", 39.3558, -9.3811),
        ("Borba", "Borba", 38.8055, -7.4546),
        ("Ferrel", "Ferrel", 39.3636, -9.3150),
        ("Coimbra", "Coimbra", 40.2033, -8.4103),
        ("Porto", "Porto", 41.1579, -8.6291),
        ("Berlengas", "Berlengas / Peniche", 39.411, -9.515),
        ("Berlengas / Peniche", "Berlengas / Peniche", 39.411, -9.515),
        ("Berlengas Peniche", "Berlengas / Peniche", 39.411, -9.515),
        ("Ilha da Berlenga", "Berlengas / Peniche", 39.411, -9.515),
        ("Ilha da Berlengas", "Berlengas / Peniche", 39.411, -9.515),
        ("Troia", "Troia", 38.4867, -8.9048),
        ("Tróia", "Troia", 38.4867, -8.9048),
    ]
    for phrase, display, lat, lon in entries:
        out[normalize_lookup(phrase)] = {"display": display, "lat": lat, "lon": lon}
    return out


REGION_FALLBACK_COORDS = _build_region_fallback_coords()


def regional_cache_query(region_key_normalized: str) -> str:
    return f"regional:{region_key_normalized}"


def region_fallback_lookup(phrase_key: str):
    """Dados de fallback para phrase_key = normalize_lookup(texto)."""
    return REGION_FALLBACK_COORDS.get(phrase_key)


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_locality(value):
    return " ".join(part.capitalize() for part in clean_text(value).split())


def normalize_city(city):
    text = clean_text(city)
    if normalize_lookup(text) in INVALID_CITY_VALUES:
        return ""
    fk = normalize_lookup(text)
    if fk in REGION_FALLBACK_COORDS:
        return REGION_FALLBACK_COORDS[fk]["display"]
    # A extração "cidade principal" fica a cargo de extract_main_city(); aqui só limpamos e capitalizamos.
    direct_key = normalize_lookup(text)
    if direct_key in CITY_CORRECTIONS:
        return CITY_CORRECTIONS[direct_key]

    words = [word for word in text.split() if normalize_lookup(word) not in REGION_WORDS]
    text = clean_text(" ".join(words))
    key = normalize_lookup(text)
    if key in CITY_CORRECTIONS:
        return CITY_CORRECTIONS[key]
    if key in INVALID_CITY_VALUES:
        return ""

    known_simple = {
        "lisboas": "Lisboa",
        "setuball": "Setúbal",
    }
    if key in known_simple:
        return known_simple[key]

    lower_words = {"de", "da", "do", "das", "dos", "e"}
    return " ".join(part.lower() if normalize_lookup(part) in lower_words else part.capitalize() for part in text.split())


def extract_main_city(raw: str) -> str:
    """Extrai a localidade mais útil para geocoding a partir de textos compostos (zonas, bairros, ilhas)."""
    text = clean_text(raw)
    if not text:
        return ""
    phrase_key = normalize_lookup(text)
    if phrase_key in LOCALITY_FALLBACK_SEQUENCES:
        return LOCALITY_FALLBACK_SEQUENCES[phrase_key][0]
    if phrase_key in REGION_FALLBACK_COORDS:
        return REGION_FALLBACK_COORDS[phrase_key]["display"]

    segments = [clean_text(s) for s in re.split(r"\s*[\,\-–—/|]\s*", text) if clean_text(s)]
    if not segments:
        return ""

    if len(segments) >= 2:
        first_key = normalize_lookup(segments[0])
        last_key = normalize_lookup(segments[-1])
        if first_key and last_key in GEO_GENERIC_REGIONS:
            return segments[0]
        if first_key and last_key not in INVALID_CITY_VALUES:
            return segments[0]

    meaningful = []
    for seg in segments:
        sk = normalize_lookup(seg)
        if sk in GEO_GENERIC_REGIONS or sk in INVALID_CITY_VALUES:
            continue
        meaningful.append(seg)

    if not meaningful:
        meaningful = segments

    chosen = meaningful[-1]

    words = [
        w
        for w in chosen.split()
        if normalize_lookup(w) not in GEO_GENERIC_REGIONS and normalize_lookup(w) not in REGION_WORDS
    ]
    chosen = clean_text(" ".join(words)) if words else chosen

    lk = normalize_lookup(chosen)
    if lk in GEO_CITY_MANUAL:
        return GEO_CITY_MANUAL[lk]

    if not lk and "algarve" in phrase_key:
        return "Algarve"
    if lk == "algarve":
        return "Algarve"

    return chosen


def geo_hint_from_postal(codigo_postal: str) -> str:
    """Referência grosseira por prefixo de código postal PT (agrupamento, não morada exata)."""
    digits = re.sub(r"\D", "", str(codigo_postal or ""))
    if len(digits) < 4:
        return ""
    try:
        prefix = int(digits[:4])
    except ValueError:
        return ""
    if 1000 <= prefix <= 1999:
        return "Lisboa"
    if 2500 <= prefix <= 2699:
        return "Oeste"
    if 2700 <= prefix <= 2799:
        return "Amadora"
    if 2800 <= prefix <= 2899:
        return "Almada"
    if prefix < 2000:
        return "Lisboa"
    if prefix < 3000:
        return "Santarem"
    if prefix < 4000:
        return "Coimbra"
    if prefix < 5000:
        return "Porto"
    if prefix < 6000:
        return "Braga"
    if prefix < 7000:
        return "Castelo Branco"
    if prefix < 8000:
        return "Evora"
    if prefix < 9000:
        return "Faro"
    return "Lisboa"


def normalize_client_type(value):
    text = normalize_lookup(value)
    if "hotel" in text:
        return "Hotel"
    if "rest" in text:
        return "Restaurante"
    return clean_text(value).title() or "Outro"


PHONE_SPLIT_RE = re.compile(r"\s*(?:/|,|;|\bou\b)\s*", re.IGNORECASE)


def normalize_phone(phone):
    digits = re.sub(r"\D", "", str(phone or ""))
    if digits.startswith("00351"):
        digits = digits[5:]
    if digits.startswith("351") and len(digits) > 9:
        digits = digits[3:]
    return digits


def extract_phone_numbers(value):
    numbers = []
    seen = set()
    for part in PHONE_SPLIT_RE.split(str(value or "")):
        normalized = normalize_phone(part)
        if normalized and normalized not in seen:
            numbers.append(normalized)
            seen.add(normalized)
    if not numbers:
        normalized = normalize_phone(value)
        if normalized:
            numbers.append(normalized)
    return numbers


def normalize_phone_list(value):
    return " / ".join(extract_phone_numbers(value))


def dedupe_fallback_key(data):
    nome = clean_text(data.get("nome_cliente") or data.get("nome_empresa") or data.get("empresa"))
    empresa = clean_text(data.get("empresa") or data.get("nome_empresa"))
    cidade = clean_text(data.get("cidade") or data.get("localidade"))
    email = clean_text(data.get("email")).lower()
    nif = normalize_phone(data.get("nif"))
    if nome and cidade:
        return f"nome_cidade:{normalize_lookup(nome)}|{normalize_lookup(cidade)}"
    if empresa and cidade:
        return f"empresa_cidade:{normalize_lookup(empresa)}|{normalize_lookup(cidade)}"
    if email:
        return f"email:{email}"
    if nif:
        return f"nif:{nif}"
    return ""


def valid_email(value):
    if not clean_text(value):
        return True
    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", clean_text(value)) is not None


def parse_float_optional(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = float(str(value).replace(",", "."))
        return parsed if math.isfinite(parsed) else None
    except ValueError:
        return None


def valid_coordinates(latitude, longitude):
    return latitude is not None and longitude is not None and -90 <= latitude <= 90 and -180 <= longitude <= 180


def parse_date(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_time_value(value):
    value = clean_text(value)
    if not value:
        return None
    match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)(?::[0-5]\d)?", value)
    if not match:
        return None
    return f"{int(match.group(1)):02d}:{match.group(2)}"


def scheduled_time_label(lead):
    return parse_time_value(getattr(lead, "hora_reuniao", None)) or ""


def scheduled_datetime_label(lead):
    if not lead or not lead.data_novo_contacto:
        return "-"
    date_text = lead.data_novo_contacto.strftime("%d/%m/%Y")
    hour_text = scheduled_time_label(lead)
    return f"{date_text} às {hour_text}" if hour_text else date_text


def scheduled_sort_key(lead):
    return (
        lead.data_novo_contacto or date.max,
        scheduled_time_label(lead) or "99:99",
        normalize_lookup(lead.nome_empresa or lead.nome_cliente or ""),
    )


def parse_split_contact_date(row):
    day = clean_text(row.get("dia"))
    month = clean_text(row.get("mes"))
    year = clean_text(row.get("ano"))
    if not (day and month and year):
        return None
    try:
        return date(int(float(year)), int(float(month)), int(float(day)))
    except (TypeError, ValueError):
        return None


def is_active_lead(lead, today=None):
    today = today or date.today()
    state = normalize_legacy_state(lead.estado)
    raw_state = normalize_lookup(lead.estado).replace(" ", "")
    if raw_state in MAP_INACTIVE_STATE_KEYS:
        return False
    if state == "Já tratado / no CRM" or state == "Sem interesse":
        return False
    if is_scheduled_lead(lead) and lead.data_novo_contacto > today:
        return False
    if lead.estado in ACTIVE_STATES:
        return True
    if lead.estado == "Adiar contacto" and lead.data_novo_contacto and lead.data_novo_contacto <= today:
        return True
    return False


def normalized_sql_text(column):
    value = db.func.lower(db.func.coalesce(column, ""))
    replacements = {
        "ã": "a",
        "á": "a",
        "à": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
        " ": "",
        "/": "",
        "-": "",
        "_": "",
        ".": "",
        ",": "",
    }
    for old, new in replacements.items():
        value = db.func.replace(value, old, new)
    return value


def aplicar_filtro_leads_operacionais_mapa(query):
    # O mapa mostra apenas leads operacionais; leads encerradas continuam disponíveis nas listagens.
    fields = [
        Lead.estado,
        Lead.estado_lead,
        Lead.classificacao_observacao,
        Lead.motivo_classificacao,
        Lead.observacoes,
        Lead.observacoes_contacto,
        Lead.tags,
    ]
    closed_conditions = []
    for field in fields:
        normalized = normalized_sql_text(field)
        closed_conditions.extend(normalized.like(f"%{token}%") for token in MAP_INACTIVE_STATE_KEYS)
    return query.filter(~or_(*closed_conditions))

def is_inactive_lead(lead, today=None):
    today = today or date.today()
    if lead.estado in {"Já tratado / no CRM", "Sem interesse", "Reunião marcada", "Sem interesse definitivo", "Cliente existente"}:
        return True
    if lead.estado == "Adiar contacto" and lead.data_novo_contacto and lead.data_novo_contacto > today:
        return True
    if lead.classificacao_observacao in {"Já tratado / no CRM", "Sem interesse", "Reunião marcada", "Sem interesse definitivo", "Cliente existente"}:
        return True
    if lead.classificacao_observacao == "Adiar contacto" and lead.data_novo_contacto and lead.data_novo_contacto > today:
        return True
    return False


def is_scheduled_lead(lead):
    if not lead or not lead.data_novo_contacto:
        return False
    state = normalize_legacy_state(lead.estado)
    text = normalize_lookup(" ".join([
        lead.estado or "",
        lead.classificacao_observacao or "",
        lead.observacoes or "",
        lead.observacoes_contacto or "",
        lead.motivo_classificacao or "",
    ]))
    if state in SCHEDULED_STATES:
        return True
    return any(token in text for token in [
        "voltar a ligar",
        "contactar mais tarde",
        "reagendada",
        "reagendar",
        "sem interesse para ja",
        "aguardar resposta",
        "ligar depois",
        "voltar a contactar",
    ])


def scheduled_bucket(lead, today=None):
    today = today or date.today()
    if not is_scheduled_lead(lead):
        return ""
    if lead.data_novo_contacto < today:
        return "overdue"
    if lead.data_novo_contacto == today:
        return "today"
    if lead.data_novo_contacto <= today + timedelta(days=7):
        return "next7"
    return "future"


def scheduled_leads_context(limit_today=None):
    today = date.today()
    buckets = {"today": [], "overdue": [], "next7": [], "future": []}
    query = Lead.query
    for lead in query.order_by(Lead.data_novo_contacto.asc(), Lead.hora_reuniao.asc(), Lead.nome_empresa.asc()).all():
        bucket = scheduled_bucket(lead, today)
        if bucket:
            buckets[bucket].append(lead)
    for items in buckets.values():
        items.sort(key=scheduled_sort_key)
    today_items = buckets["today"][:limit_today] if limit_today else buckets["today"]
    return {
        "today": today_items,
        "overdue": buckets["overdue"],
        "next7": buckets["next7"],
        "future": buckets["future"],
        "today_count": len(buckets["today"]),
        "overdue_count": len(buckets["overdue"]),
        "badge_count": len(buckets["today"]) + len(buckets["overdue"]),
    }


def scheduled_badge_count():
    today = date.today()
    return (
        Lead.query.filter(
            Lead.data_novo_contacto.isnot(None),
            Lead.data_novo_contacto <= today,
            Lead.estado.in_(list(SCHEDULED_STATES)),
        ).count()
    )


def operational_notifications_context(limit=8):
    today = date.today()
    if not getattr(current_user, "is_authenticated", False):
        return {"count": 0, "items": []}

    followups_today = [
        lead
        for lead in Lead.query.filter(Lead.data_novo_contacto == today)
        .order_by(Lead.hora_reuniao.asc(), Lead.nome_empresa.asc())
        .limit(30)
        .all()
        if is_scheduled_lead(lead)
    ]
    overdue = [
        lead
        for lead in Lead.query.filter(
            Lead.data_novo_contacto.isnot(None),
            Lead.data_novo_contacto < today,
        )
        .order_by(Lead.data_novo_contacto.asc(), Lead.hora_reuniao.asc(), Lead.nome_empresa.asc())
        .limit(30)
        .all()
        if is_scheduled_lead(lead)
    ]
    meetings_today = (
        Lead.query.filter(Lead.data_reuniao == today)
        .order_by(Lead.hora_reuniao.asc(), Lead.nome_empresa.asc())
        .limit(30)
        .all()
    )

    items = []

    def lead_name(lead):
        return lead.nome_cliente or lead.nome_empresa or lead.empresa or "Lead sem nome"

    def push(lead, kind, action, when_label):
        items.append({
            "kind": kind,
            "time": scheduled_time_label(lead) or when_label,
            "lead": lead_name(lead),
            "action": action,
            "url": url_for("mapa_leads", lead_id=lead.id),
        })

    for lead in followups_today:
        push(lead, "today", "Follow-up", "Hoje")
    for lead in meetings_today:
        push(lead, "meeting", "Reunião", "Hoje")
    for lead in overdue:
        days = max(1, (today - lead.data_novo_contacto).days)
        push(lead, "overdue", "Atrasada", f"{days}d")

    priority = {"overdue": 0, "today": 1, "meeting": 2}
    items.sort(key=lambda item: (priority.get(item["kind"], 9), item["time"] or "99:99", normalize_lookup(item["lead"])))
    return {
        "count": len(followups_today) + len(meetings_today) + len(overdue),
        "items": items[:limit],
    }


def classify_observations(observacoes="", observacoes_contacto=""):
    combined = clean_text(f"{observacoes} {observacoes_contacto}")
    if not combined:
        return "Por contactar", "", ""
    text = normalize_lookup(combined)
    text = (
        text.replace("reuni?o", "reuniao")
        .replace("observa??o", "observacao")
        .replace("n?o", "nao")
        .replace("j?", "ja")
        .replace("poss?vel", "possivel")
    )

    def log(pattern, state):
        print(f"[CLASSIFY] observacao='{combined}' padrao='{pattern or '-'}' estado='{state}'", flush=True)

    def contains_any(patterns):
        for pattern in patterns:
            if pattern in text:
                return pattern
        return ""

    meeting_ambiguous = [
        "marcar reuniao", "agendar reuniao", "tentar marcar", "ligar para marcar",
        "possivel reuniao", "possível reunião", "quer reuniao", "falar sobre reuniao",
        "voltar a falar",
    ]
    if contains_any(meeting_ambiguous):
        log(contains_any(meeting_ambiguous), "Por contactar")
        return "Por contactar", "", "Expressao ambigua; necessita contacto antes de assumir estado."

    rules = [
        ("Já tratado / no CRM", [
            "ja cliente", "já cliente", "cliente existente", "ja tem sistema", "já tem sistema",
            "ja trabalha connosco", "já trabalha connosco",
        ], "Padrao forte de cliente existente."),
        ("Já tratado / no CRM", [
            "reuniao marcada", "reunião marcada", "reuniao agendada", "reunião agendada",
            "agendado para", "visita marcada", "confirmada reuniao", "confirmada reunião",
            "reuniao dia", "reunião dia", "reuniao as", "reunião as", "reuniao às", "reunião às",
            "marcado para dia", "ficou marcado", "encontro confirmado",
        ], "Padrao forte de tratamento externo no CRM."),
        ("Sem interesse", [
            "nao interessado", "não interessado", "sem interesse", "nao quer", "não quer",
            "recusou", "dispensou", "nao pretende", "não pretende",
            "nao tem interesse", "não tem interesse", "cliente nao quer", "cliente não quer",
        ], "Padrao forte de falta de interesse definitiva."),
        ("Ligar de volta", [
            "ligar depois", "ligar mais tarde", "voltar a ligar", "voltar a contactar",
            "daqui a uns meses", "nao agora", "não agora", "falar mais tarde",
            "ligar semana que vem", "talvez mais tarde", "contactar depois",
            "contactar mais tarde", "sem resposta", "nao atendeu", "não atendeu",
            "nao respondeu", "não respondeu",
        ], "Padrao de retoma futura de contacto."),
    ]

    for state, patterns, reason in rules:
        pattern = contains_any(patterns)
        if pattern:
            log(pattern, state)
            return state, state, f"{reason} Padrao detetado: '{pattern}'."

    log("", "Por contactar")
    return "Por contactar", "", "Sem padrao forte; mantido conservadoramente por contactar."


def current_audit_user_id():
    if has_request_context() and current_user.is_authenticated:
        return current_user.id
    return getattr(AUDIT_CONTEXT, "user_id", None)


def infer_action_type(action):
    value = normalize_lookup(action)
    if "import" in value:
        return "import_excel"
    if "adicion" in value or "criad" in value:
        return "criacao_lead"
    if "reuniao" in value or "crm" in value:
        return "reuniao_marcada"
    if "reagend" in value or "adiar" in value or "follow" in value:
        return "followup_reagendado"
    if "nota" in value:
        return "nota_adicionada"
    if "apag" in value or "elimin" in value:
        return "apagar_lead"
    if "contact" in value or "atualiz" in value or "corrig" in value or "alterad" in value or "atribuid" in value or "removid" in value:
        return "edicao_lead"
    return "atividade"


def registar_historico(lead, action, observation="", commercial=None, tipo_acao=None, resultado=None, created_at=None, user_id=None):
    db.session.add(HistoricoLead(
        lead=lead,
        user_id=current_audit_user_id() if user_id is None else user_id,
        acao=action,
        tipo_acao=tipo_acao or infer_action_type(action),
        observacao=observation,
        resultado=resultado or "ok",
        comercial_responsavel=commercial or lead.comercial_responsavel,
        created_at=created_at or datetime.utcnow(),
    ))


def add_history(lead, action, observation="", commercial=None, tipo_acao=None, resultado=None, created_at=None, user_id=None):
    registar_historico(
        lead=lead,
        action=action,
        observation=observation,
        commercial=commercial,
        tipo_acao=tipo_acao,
        resultado=resultado,
        created_at=created_at,
        user_id=user_id,
    )


def read_csv(file_storage):
    content = file_storage.stream.read()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            except csv.Error:
                dialect = csv.excel
            return list(csv.DictReader(StringIO(text), dialect=dialect))
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao foi possivel ler o CSV. Guarda o ficheiro como UTF-8 ou Excel .xlsx.")


HEADER_KEYWORDS = {"nome", "morada", "contacto", "telefone", "telemovel", "localidade", "cidade", "obs", "observacoes", "codigo", "cod", "cp"}
TECHNICAL_SHEET_WORDS = {"config", "tabela", "pivot", "resumo", "grafico", "gráfico", "dashboard", "template", "lista"}


def row_has_content(row):
    return any(clean_text(value) for value in row or [])


def is_relevant_sheet(sheet):
    if sheet.max_row < 2 or sheet.max_column < 2:
        return False
    title = normalize_lookup(sheet.title)
    if any(word in title for word in TECHNICAL_SHEET_WORDS) and not any(word in title for word in {"cliente", "pendente", "horeca", "venda"}):
        return False
    sample_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True))
    return any(row_has_content(row) for row in sample_rows)


def header_score(values):
    normalized = [normalize_lookup(value) for value in values if clean_text(value)]
    if len(normalized) < 2:
        return 0
    score = 0
    for item in normalized:
        compact = item.replace(" ", "")
        if any(keyword in item or keyword in compact for keyword in HEADER_KEYWORDS):
            score += 1
    return score


def detect_header_row(sheet):
    best_index = None
    best_values = []
    best_score = 0
    max_scan = min(sheet.max_row, 25)
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        score = header_score(row)
        if score > best_score:
            best_index = index
            best_values = row
            best_score = score
    if best_score < 2:
        return None, []
    headers = []
    seen = {}
    for idx, value in enumerate(best_values, start=1):
        key = normalize_key(value) if clean_text(value) else f"coluna_{idx}"
        seen[key] = seen.get(key, 0) + 1
        headers.append(key if seen[key] == 1 else f"{key}_{seen[key]}")
    return best_index, headers


def is_technical_row(values):
    text = " ".join(normalize_lookup(value) for value in values.values() if clean_text(value))
    if not text:
        return True
    technical_terms = ["total", "subtotal", "pagina", "página", "observacoes gerais", "cabecalho", "cabeçalho"]
    useful_terms = ["rua", "avenida", "av ", "telefone", "@", "lda", "restaurante", "hotel"]
    if any(term in text for term in useful_terms):
        return False
    return len(text) < 4 or any(text == term for term in technical_terms)


def get_cell_value(cell):
    return cell.value if hasattr(cell, "value") else cell


def read_xlsx(file_storage):
    # Fluxo estabilizado (Excel principal Alltera):
    # - apenas a primeira worksheet
    # - primeira linha como header
    # - sem heurísticas pesadas
    workbook = load_workbook(file_storage.stream, data_only=True)
    if not workbook.worksheets:
        return []
    sheet = workbook.worksheets[0]

    # Importante: iter_rows(values_only=True) já devolve valores “crus”.
    # Aqui usamos values_only=False para manter compatibilidade com cell.value.
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = [normalize_key(get_cell_value(c)) for c in header_cells]

    rows = []
    for row_cells in sheet.iter_rows(min_row=2, values_only=False):
        row_values = [get_cell_value(c) for c in row_cells]
        if not row_has_content(row_values):
            continue
        values = {headers[idx]: row_values[idx] for idx in range(min(len(headers), len(row_values))) if headers[idx]}
        rows.append(values)
    return rows


def read_import_file(file_storage):
    filename = file_storage.filename.lower()
    if filename.endswith(".csv"):
        return read_csv(file_storage)
    if filename.endswith(".xlsx"):
        return read_xlsx(file_storage)
    raise ValueError("Formato invalido. Usa .xlsx ou .csv.")


def lead_value(lead, field):
    if isinstance(lead, dict):
        return clean_text(lead.get(field))
    return clean_text(getattr(lead, field, ""))


def geocoding_queries_for_lead(lead):
    """Lista de consultas Nominatim por ordem de preferência (aproximação geográfica)."""
    raw = lead_value(lead, "cidade") or lead_value(lead, "localidade")
    morada = lead_value(lead, "morada")
    extracted = extract_main_city(raw)
    cp = lead_value(lead, "codigo_postal")

    candidates = []
    seen = set()

    def add_query(query):
        query = clean_text(query)
        if not query or query in seen:
            return
        seen.add(query)
        candidates.append(query)

    def add_place(name):
        place = normalize_city(name)
        if not place:
            return
        add_query(f"{place}, Portugal")

    if morada and cp and raw:
        add_query(f"{morada}, {cp}, {raw}, Portugal")
    if morada and raw:
        add_query(f"{morada}, {raw}, Portugal")
    if cp and raw:
        add_query(f"{cp}, {raw}, Portugal")
    if cp:
        add_query(f"{cp}, Portugal")

    add_place(extracted)
    add_place(raw)

    hint = geo_hint_from_postal(cp)
    if hint and hint not in seen:
        add_place(hint)

    return candidates


def geocode_lead(lead, import_cache=None):
    latitude = parse_float_optional(lead_value(lead, "latitude"))
    longitude = parse_float_optional(lead_value(lead, "longitude"))
    if valid_coordinates(latitude, longitude):
        return {
            "latitude": latitude,
            "longitude": longitude,
            "query_usada": "",
            "sucesso": True,
            "erro": "",
            "tipo": None,
            "display_regiao": None,
            "cache": False,
        }

    import_cache = import_cache if import_cache is not None else {}
    last_error = {
        "latitude": None,
        "longitude": None,
        "query_usada": "",
        "sucesso": False,
        "erro": "Sem resultado para a morada/localidade.",
        "tipo": None,
        "display_regiao": None,
        "cache": False,
    }

    raw = clean_text(lead_value(lead, "cidade") or lead_value(lead, "localidade"))
    fk = normalize_lookup(raw) if raw else ""
    algarve_key = normalize_lookup("Algarve")

    def apply_region_fallback(region_norm_key: str, modo: str):
        fb = region_fallback_lookup(region_norm_key)
        if not fb:
            return None
        q = regional_cache_query(region_norm_key)
        if q in import_cache:
            cached_result = import_cache[q]
            if cached_result.get("sucesso"):
                return {
                    **cached_result,
                    "query_usada": q,
                    "tipo": "fallback_regiao",
                    "display_regiao": fb["display"],
                    "modo_fallback": modo,
                    "cache": True,
                }

        with db.session.no_autoflush:
            cached = GeocodingCache.query.filter_by(query_text=q).first()
        if cached and valid_coordinates(cached.latitude, cached.longitude):
            result = {
                "latitude": cached.latitude,
                "longitude": cached.longitude,
                "query_usada": q,
                "sucesso": True,
                "erro": "",
                "tipo": "fallback_regiao",
                "display_regiao": fb["display"],
                "modo_fallback": modo,
                "cache": True,
            }
            import_cache[q] = result
            return result

        lat, lon = fb["lat"], fb["lon"]
        if cached:
            cached.latitude = lat
            cached.longitude = lon
        else:
            # UNIQUE(query_text): pode ser inserido por outra linha do mesmo import.
            # Antes de add() fazemos lookup para não rebentar com IntegrityError.
            existing = GeocodingCache.query.filter_by(query_text=q).first()
            if existing:
                existing.latitude = lat
                existing.longitude = lon
            else:
                safe_cache_geocoding(q, lat, lon)
        result = {
            "latitude": lat,
            "longitude": lon,
            "query_usada": q,
            "sucesso": True,
            "erro": "",
            "tipo": "fallback_regiao",
            "display_regiao": fb["display"],
            "modo_fallback": modo,
            "cache": False,
        }
        import_cache[q] = result
        return result

    # 1) Célula inteira = região conhecida (sem Nominatim)
    if fk and fk in REGION_FALLBACK_COORDS:
        r = apply_region_fallback(fk, "celula")
        if r:
            return r

    # 2) Nominatim para localidades específicas (ex.: Ferreiras após "Algarve - Ferreiras")
    for query in geocoding_queries_for_lead(lead):
        if not query:
            continue
        if query in import_cache:
            cached_result = import_cache[query]
            if cached_result.get("sucesso"):
                return {**cached_result, "query_usada": query, "cache": True}
            last_error = {**cached_result, "query_usada": query, "cache": True}
            continue

        with db.session.no_autoflush:
            cached = GeocodingCache.query.filter_by(query_text=query).first()
        if cached:
            cached_success = valid_coordinates(cached.latitude, cached.longitude)
            result = {
                "latitude": cached.latitude,
                "longitude": cached.longitude,
                "query_usada": query,
                "sucesso": cached_success,
                "erro": "" if cached_success else "Sem resultado em cache para esta pesquisa.",
                "tipo": None,
                "display_regiao": None,
                "cache": True,
            }
            import_cache[query] = result
            if cached_success:
                return result
            last_error = result
            continue

        try:
# pequenas pausas evitam sobrecarregar o Nominatim
            # (no modo estabilizado vamos limitar as chamadas e o timeout)
            response = requests.get(
                NOMINATIM_URL,
                params={"q": query, "format": "json", "limit": 1, "countrycodes": "pt"},
                headers={"User-Agent": NOMINATIM_USER_AGENT},
                timeout=5,
            )
            response.raise_for_status()
            payload = response.json()
            if payload:
                lat = parse_float_optional(payload[0].get("lat"))
                lon = parse_float_optional(payload[0].get("lon"))
                if valid_coordinates(lat, lon):
                    safe_cache_geocoding(query, lat, lon)
                    result = {
                        "latitude": lat,
                        "longitude": lon,
                        "query_usada": query,
                        "sucesso": True,
                        "erro": "",
                        "tipo": None,
                        "display_regiao": None,
                        "cache": False,
                    }
                    import_cache[query] = result
                    return result
            safe_cache_geocoding(query, None, None)
            result = {
                "latitude": None,
                "longitude": None,
                "query_usada": query,
                "sucesso": False,
                "erro": "Sem resultado para esta cidade.",
                "tipo": None,
                "display_regiao": None,
                "cache": False,
            }
            import_cache[query] = result
            last_error = result
        except Exception as exc:
            result = {
                "latitude": None,
                "longitude": None,
                "query_usada": query,
                "sucesso": False,
                "erro": str(exc),
                "tipo": None,
                "display_regiao": None,
                "cache": False,
            }
            import_cache[query] = result
            last_error = result

    # 3) Texto continha "Algarve" mas a localidade específica não geocodificou — centro aproximado do Algarve
    if fk and "algarve" in fk and region_fallback_lookup(algarve_key):
        ext = extract_main_city(raw)
        ext_k = normalize_lookup(ext) if ext else ""
        if ext_k != algarve_key and ext_k not in REGION_FALLBACK_COORDS:
            r = apply_region_fallback(algarve_key, "apos_falha_geocoding")
            if r:
                return r

    return last_error


def detect_column_mapping(headers):
    mapping = {}
    normalized_headers = {normalize_key(header): header for header in headers}
    compact_headers = {compact_key(header): header for header in headers}
    used_headers = set()
    for field, aliases in FIELD_ALIASES.items():
        detected = ""
        for alias in aliases:
            key = normalize_key(alias)
            compact = compact_key(alias)
            if key in normalized_headers and normalized_headers[key] not in used_headers:
                detected = normalized_headers[key]
                break
            if compact in compact_headers and compact_headers[compact] not in used_headers:
                detected = compact_headers[compact]
                break
        if not detected:
            best_header = ""
            best_score = 0
            alias_forms = [compact_key(alias) for alias in aliases if compact_key(alias)]
            for header in headers:
                if header in used_headers:
                    continue
                header_form = compact_key(header)
                for alias_form in alias_forms:
                    score = SequenceMatcher(None, header_form, alias_form).ratio()
                    if score > best_score:
                        best_score = score
                        best_header = header
            if best_score >= 0.70:
                detected = best_header
        if detected:
            used_headers.add(detected)
        mapping[field] = detected
    return mapping


def apply_mapping(rows, mapping):
    normalized_rows = []
    for row in rows:
        row_mapping = mapping
        if any(source and source not in row for source in mapping.values()):
            row_mapping = detect_column_mapping(list(row.keys()))
        normalized = {}
        for target, source in row_mapping.items():
            normalized[target] = row.get(source, "") if source else ""
        for key, value in row.items():
            if str(key).startswith("_"):
                normalized[key] = value
        normalized_rows.append(normalized)
    return normalized_rows


def similarity(a, b):
    return SequenceMatcher(None, normalize_lookup(a), normalize_lookup(b)).ratio()


def find_duplicate(candidate):
    with db.session.no_autoflush:
        candidate_phones = extract_phone_numbers(candidate.get("telefone"))
        candidate_fallback = dedupe_fallback_key(candidate) if not candidate_phones else ""
        leads = Lead.query.all()
        for lead in leads:
            lead_phones = extract_phone_numbers(lead.telefone)
            matching_phones = sorted(set(candidate_phones) & set(lead_phones))
            if matching_phones:
                return lead, f"Contacto telefonico normalizado: {', '.join(matching_phones)}"

            if candidate_fallback and candidate_fallback == dedupe_fallback_key(lead.to_dict()):
                return lead, "Fallback sem telefone: nome/empresa + cidade, email ou NIF"

            same_postal = bool(candidate.get("codigo_postal") and lead.codigo_postal and normalize_lookup(lead.codigo_postal) == normalize_lookup(candidate["codigo_postal"]))
            address_close = bool(candidate.get("morada") and similarity(lead.morada or "", candidate["morada"]) >= 0.82)
            name_close = similarity(lead.nome_empresa, candidate["nome_empresa"]) >= 0.82
            if same_postal and address_close and name_close:
                return lead, "Nome + morada + codigo postal semelhantes"
    return None, None


def find_possible_duplicate(candidate):
    for lead in Lead.query.all():
        same_locality = normalize_lookup(lead.localidade) == normalize_lookup(candidate["localidade"])
        name_close = max(
            similarity(lead.nome_empresa, candidate["nome_empresa"]),
            similarity(lead.nome_cliente or "", candidate.get("nome_cliente") or ""),
            similarity(lead.empresa or "", candidate.get("empresa") or ""),
        ) >= 0.72
        lead_phones = extract_phone_numbers(lead.telefone or "")
        candidate_phones = extract_phone_numbers(candidate.get("telefone"))
        phone_close = any(
            lead_phone and candidate_phone and lead_phone[-7:] == candidate_phone[-7:]
            for lead_phone in lead_phones
            for candidate_phone in candidate_phones
        )
        email_close = bool(candidate.get("email") and lead.email and similarity(lead.email, candidate["email"]) >= 0.88)
        nif_equal = bool(candidate.get("nif") and lead.nif and normalize_lookup(lead.nif) == normalize_lookup(candidate["nif"]))
        address_close = bool(candidate["morada"] and similarity(lead.morada or "", candidate["morada"]) >= 0.70)
        if nif_equal:
            return lead, "NIF igual"
        if email_close and name_close:
            return lead, "email e nome semelhantes"
        if same_locality and name_close and (phone_close or address_close or similarity(lead.nome_empresa, candidate["nome_empresa"]) >= 0.82):
            return lead, "Possível duplicado detectado: nome/localidade semelhantes com telefone, email ou morada aproximados"
    return None, None


def update_empty_fields(existing, candidate):
    updated = False
    for field, value in candidate.items():
        if field in {"observacoes", "observacoes_contacto"}:
            current = clean_text(getattr(existing, field))
            incoming = clean_text(value)
            if incoming and not current:
                setattr(existing, field, incoming)
                updated = True
            elif incoming and incoming not in current:
                setattr(existing, field, f"{current}\n{incoming}")
                updated = True
            continue
        if field in {"latitude", "longitude"}:
            if getattr(existing, field) is None and value is not None:
                setattr(existing, field, value)
                updated = True
        elif hasattr(existing, field) and not clean_text(getattr(existing, field)) and clean_text(value):
            setattr(existing, field, value)
            updated = True
    return updated


def validate_and_import_rows(rows, auto_geocode=True, geocode_city_limit=None):
    geocode_city_limit = int(geocode_city_limit) if geocode_city_limit else None
    summary = {
        "read": len(rows),
        "imported": 0,
        "duplicates": 0,
        "updated": 0,
        "errors": [],
        "without_coordinates": 0,
        "geocoded": 0,
        "geocoding_errors": [],
        "cities_geocoded": [],
        "cities_not_found": [],
        "city_normalization": [],
        "classification_counts": {
            "Ligar de volta": 0,
            "Adiar contacto": 0,
            "Sem interesse definitivo": 0,
            "Reuni?o marcada": 0,
            "Rever manualmente": 0,
        },
        "sheets_imported": sorted({clean_text(row.get("_sheet_name")) for row in rows if clean_text(row.get("_sheet_name"))}),
    }
    geocoded_cities = set()
    not_found_cities = set()
    city_normalization = {}
    import_geocode_cache = {}

    for index, raw_row in enumerate(rows, start=2):
        try:
            row = {normalize_key(key): value for key, value in raw_row.items()}
            if not any(clean_text(value) for key, value in row.items() if not str(key).startswith("_")):
                continue

            nome_cliente = clean_text(row.get("contacto") or row.get("nome_cliente"))
            empresa = clean_text(row.get("nome_empresa") or row.get("empresa"))
            area_negocio = clean_text(row.get("tipo_cliente") or row.get("area_negocio") or row.get("categoria")) or "Outro"
            cidade_original = clean_text(row.get("localidade") or row.get("cidade"))
            extracted = extract_main_city(cidade_original)
            norm_extract = normalize_city(extracted) if extracted else ""

            if cidade_original and normalize_lookup(cidade_original) not in INVALID_CITY_VALUES:
                display_city = normalize_locality(cidade_original)
            else:
                display_city = "Sem cidade"

            nome_principal = nome_cliente or empresa or "Sem nome"
            candidate = {
                "nome_cliente": nome_principal,
                "area_negocio": area_negocio,
                "cidade": display_city,
                "empresa": empresa,
                "nome_empresa": nome_principal,
                "tipo_cliente": area_negocio,
                "morada": clean_text(row.get("morada")),
                "codigo_postal": clean_text(row.get("codigo_postal")),
                "localidade": display_city,
                "contacto": nome_principal,
                "telefone": normalize_phone_list(row.get("telefone")),
                "email": clean_text(row.get("email")).lower(),
                "comercial_responsavel": clean_text(row.get("comercial_responsavel")) or UNASSIGNED_COMMERCIAL,
                "categoria": clean_text(row.get("categoria")),
                "nif": clean_text(row.get("nif")),
                "observacoes": clean_text(row.get("observacoes")),
                "observacoes_contacto": clean_text(row.get("observacoes_contacto")),
                "reuniao_info": clean_text(row.get("reuniao_info")),
                "latitude": parse_float_optional(row.get("latitude")),
                "longitude": parse_float_optional(row.get("longitude")),
            }
            print(
                f"[IMPORT] linha {index}: telefone normalizado='{candidate['telefone'] or '-'}' "
                f"fallback='{dedupe_fallback_key(candidate) or '-'}'",
                flush=True,
            )

            estado_sugerido, classificacao, motivo = classify_observations(candidate["observacoes"], candidate["observacoes_contacto"])
            candidate["classificacao_observacao"] = classificacao
            candidate["motivo_classificacao"] = motivo
            if estado_sugerido == "Adiar contacto":
                candidate["data_novo_contacto"] = date.today() + timedelta(days=30)
            contact_date = parse_split_contact_date(row)
            if classificacao in summary["classification_counts"]:
                summary["classification_counts"][classificacao] += 1

            geocode_result = None
            if auto_geocode and display_city != "Sem cidade" and not valid_coordinates(candidate["latitude"], candidate["longitude"]):
                geocode_result = geocode_lead(candidate, import_cache=import_geocode_cache)
                if geocode_result["sucesso"]:
                    candidate["latitude"] = geocode_result["latitude"]
                    candidate["longitude"] = geocode_result["longitude"]
                    summary["geocoded"] += 1
                    geocoded_cities.add(norm_extract or extracted or display_city)
                else:
                    not_found_cities.add(norm_extract or extracted or display_city)
                    summary["geocoding_errors"].append({
                        "line": index,
                        "name": candidate["nome_cliente"],
                        "query": geocode_result.get("query_usada", ""),
                        "error": geocode_result.get("erro", "Erro desconhecido"),
                    })

            if not valid_coordinates(candidate["latitude"], candidate["longitude"]):
                summary["without_coordinates"] += 1

            city_key = (cidade_original or "-", extracted or "-", norm_extract or "-")
            if city_key not in city_normalization:
                if valid_coordinates(candidate["latitude"], candidate["longitude"]):
                    city_status = "sucesso"
                    if geocode_result and geocode_result.get("sucesso"):
                        geo_msg = f"Sucesso ? {geocode_result.get('query_usada', '')}"
                    else:
                        geo_msg = "Sucesso (coordenadas dispon?veis)"
                    cadeia = f"{cidade_original or '-'} -> {norm_extract or extracted or '-'} -> sucesso"
                elif display_city == "Sem cidade":
                    city_status = "ignorada"
                    geo_msg = "Ignorada (sem cidade ?til no Excel)"
                    cadeia = f"{cidade_original or '-'} -> ? -> ignorada"
                else:
                    city_status = "falha"
                    geo_msg = f"Falha ? ?ltima query: {geocode_result.get('query_usada', '-') if geocode_result else '-'}"
                    cadeia = f"{cidade_original or '-'} -> {norm_extract or extracted or '-'} -> falha"
                city_normalization[city_key] = {
                    "original": cidade_original or "-",
                    "extraida": extracted or "-",
                    "normalizada": norm_extract or "-",
                    "geocoding": geo_msg,
                    "resultado": city_status,
                    "cadeia": cadeia,
                }

            duplicate, reason = find_duplicate(candidate)
            if duplicate:
                summary["duplicates"] += 1
                print(f"[IMPORT] linha {index}: duplicado encontrado lead_id={duplicate.id} motivo={reason}", flush=True)
                if update_empty_fields(duplicate, candidate):
                    summary["updated"] += 1
                    add_history(
                        duplicate,
                        "Lead atualizada por duplicado",
                        "Lead duplicada encontrada na importação; dados atualizados.",
                    )
                    print(f"[IMPORT] linha {index}: lead atualizada lead_id={duplicate.id}", flush=True)
                if classificacao in summary["classification_counts"]:
                    summary["classification_counts"][classificacao] -= 1
                add_history(
                    duplicate,
                    "Duplicado encontrado",
                    f"Lead duplicada encontrada na importação; dados atualizados. Motivo: {reason}.",
                )
                if contact_date:
                    add_history(
                        duplicate,
                        "Contacto registado em importacao",
                        "Data reconstruida a partir das colunas DIA/MES/ANO.",
                        created_at=datetime.combine(contact_date, datetime.min.time()),
                        tipo_acao="import_excel",
                    )
                db.session.commit()
                continue

            possible_duplicate, possible_reason = find_possible_duplicate(candidate)
            if possible_duplicate:
                db.session.add(PossivelDuplicado(
                    lead=possible_duplicate,
                    dados_importados=json.dumps(candidate, default=str, ensure_ascii=False),
                    motivo=possible_reason,
                ))
                add_history(possible_duplicate, "Poss?vel duplicado detectado", possible_reason)

            lead = Lead(**candidate, estado=estado_sugerido, estado_lead=estado_sugerido)
            db.session.add(lead)
            db.session.flush()
            print(f"[IMPORT] linha {index}: nova lead criada lead_id={lead.id}", flush=True)
            add_history(lead, "Lead importada", "Lead criada por importacao de ficheiro.")
            if contact_date:
                add_history(
                    lead,
                    "Contacto registado em importacao",
                    "Data reconstruida a partir das colunas DIA/MES/ANO.",
                    created_at=datetime.combine(contact_date, datetime.min.time()),
                    tipo_acao="import_excel",
                )
            if classificacao:
                add_history(lead, "Classificacao automatica", f"{classificacao}. {motivo}")
            if geocode_result:
                if geocode_result["sucesso"]:
                    add_history(lead, "geocoding autom?tico", f"Query: {geocode_result['query_usada']}. Resultado: {candidate['latitude']}, {candidate['longitude']}.")
                else:
                    add_history(lead, "Erro de geocoding autom?tico", f"Query: {geocode_result.get('query_usada', '')}. Erro: {geocode_result.get('erro', '')}.")
            summary["imported"] += 1
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            summary["errors"].append({"line": index, "error": str(exc)})
            print(f"[IMPORT] erro na linha {index}: {exc}", flush=True)
            continue

    summary["cities_geocoded"] = sorted(geocoded_cities)
    summary["cities_not_found"] = sorted(not_found_cities)
    summary["city_normalization"] = list(city_normalization.values())
    return summary


# Motor de importacao limpo e previsivel. As definicoes abaixo substituem a
# pipeline antiga sem alterar o mapa, planeamento ou templates principais.
IMPORT_FIELD_ALIASES = {
    "comercial_responsavel": ["comercial", "comercial responsavel", "comercial_responsavel"],
    "categoria": ["categoria"],
    "tipo_cliente": ["area de negocio", "area negocio", "area_de_negocio", "área de negócio", "tipo cliente", "tipo", "segmento"],
    "nome_cliente": ["nome cliente", "nome_cliente", "nome", "cliente"],
    "telefone": ["contacto telefonico", "contacto telefónico", "contacto", "telefone", "telemovel", "telemóvel", "tlm", "tel"],
    "email": ["email", "e-mail", "mail"],
    "nome_empresa": ["empresa", "nome empresa", "nome_empresa", "designacao", "designação"],
    "nif": ["nif", "numero fiscal", "número fiscal"],
    "cidade": ["cidade", "localidade", "concelho", "zona"],
    "localidade": ["cidade", "localidade", "concelho", "zona"],
    "morada": ["morada", "endereco", "endereço", "rua"],
    "codigo_postal": ["cod p", "cod. p.", "codigo postal", "código postal", "codigo_postal", "cp", "c.p."],
    "observacoes": ["observacoes", "observações", "obs", "notas"],
    "observacoes_contacto": ["observacoes do contacto", "observações do contacto", "observacoes_contacto", "notas contacto"],
    "reuniao_info": ["reuniao", "reunião"],
    "latitude": ["latitude", "lat"],
    "longitude": ["longitude", "lng", "lon"],
    "dia": ["dia"],
    "mes": ["mes", "mês"],
    "ano": ["ano"],
}


def import_header_key(value):
    text = normalize_lookup(value)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return clean_text(text)


def normalize_headers(headers):
    normalized = []
    seen = {}
    for index, header in enumerate(headers, start=1):
        key = import_header_key(header) or f"coluna {index}"
        seen[key] = seen.get(key, 0) + 1
        normalized.append(key if seen[key] == 1 else f"{key} {seen[key]}")
    return normalized


def map_columns(headers):
    mapping = {}
    normalized_headers = {import_header_key(header): header for header in headers if clean_text(header)}
    compact_headers = {re.sub(r"[^a-z0-9]", "", key): header for key, header in normalized_headers.items()}
    used = set()
    for field, aliases in IMPORT_FIELD_ALIASES.items():
        detected = ""
        for alias in aliases:
            alias_key = import_header_key(alias)
            alias_compact = re.sub(r"[^a-z0-9]", "", alias_key)
            if alias_key in normalized_headers and normalized_headers[alias_key] not in used:
                detected = normalized_headers[alias_key]
                break
            if alias_compact in compact_headers and compact_headers[alias_compact] not in used:
                detected = compact_headers[alias_compact]
                break
        if detected:
            mapping[field] = detected
            used.add(detected)
        else:
            best_header = ""
            best_score = 0
            alias_forms = [re.sub(r"[^a-z0-9]", "", import_header_key(alias)) for alias in aliases]
            for header in headers:
                if header in used:
                    continue
                header_form = re.sub(r"[^a-z0-9]", "", import_header_key(header))
                if not header_form:
                    continue
                for alias_form in alias_forms:
                    score = SequenceMatcher(None, header_form, alias_form).ratio()
                    if score > best_score:
                        best_score = score
                        best_header = header
            if best_score >= 0.82:
                mapping[field] = best_header
                used.add(best_header)
            else:
                mapping[field] = ""
    return mapping


def read_import_file(file_storage, sheet_name=None):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        content = file_storage.stream.read()
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = content.decode(encoding)
                sample = text[:2048]
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;")
                except csv.Error:
                    dialect = csv.excel
                rows = list(csv.DictReader(StringIO(text), dialect=dialect))
                return [row for row in rows if row_has_content(row.values())]
            except UnicodeDecodeError:
                continue
        raise ValueError("Nao foi possivel ler o CSV. Guarda o ficheiro como UTF-8 ou Excel .xlsx.")

    if filename.endswith(".xlsx"):
        workbook = load_workbook(file_storage.stream, data_only=True, read_only=True)
        if not workbook.worksheets:
            return []
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("A folha selecionada nao existe no ficheiro.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.worksheets[0]
        if sheet.max_row < 2 or sheet.max_column < 2:
            raise ValueError("A folha selecionada esta vazia ou tem menos de 2 colunas.")
        raw_rows = list(sheet.iter_rows(values_only=True))
        header_index = None
        headers = []
        best_score = 0
        for index, row in enumerate(raw_rows[:20]):
            candidate_headers = [clean_text(value) for value in row]
            mapping = map_columns(candidate_headers)
            score = sum(1 for value in mapping.values() if value)
            if score > best_score:
                best_score = score
                header_index = index
                headers = candidate_headers
        if header_index is None or best_score < 2:
            raise ValueError("Nao foi possivel detetar cabecalhos no ficheiro.")

        rows = []
        for row in raw_rows[header_index + 1:]:
            if not row_has_content(row):
                continue
            item = {}
            for index, header in enumerate(headers):
                if header:
                    item[header] = row[index] if index < len(row) else ""
            rows.append(item)
        return rows

    raise ValueError("Formato invalido. Usa .xlsx ou .csv.")


def sheet_kind(name):
    key = normalize_lookup(name)
    if any(word in key for word in ["leads", "clientes", "horeca", "vendas", "contactos"]):
        return "recommended"
    if any(word in key for word in ["manutencao", "manutenção", "contratos", "assistencia", "assistência"]):
        return "warning"
    return "neutral"


def list_import_sheets(file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        preview_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True))
        header_index = None
        headers = []
        best_score = 0
        for index, row in enumerate(preview_rows):
            candidate_headers = [clean_text(value) for value in row]
            score = sum(1 for value in map_columns(candidate_headers).values() if value)
            if score > best_score:
                best_score = score
                header_index = index
                headers = candidate_headers
        columns = [header for header in headers if header][:6]
        approx_rows = max((sheet.max_row or 0) - ((header_index or 0) + 1), 0) if sheet.max_row else 0
        sheets.append({
            "name": sheet.title,
            "rows": approx_rows,
            "columns": columns,
            "kind": sheet_kind(sheet.title),
            "valid": sheet.max_row >= 2 and sheet.max_column >= 2,
        })
    return sheets


def parse_rows(rows, mapping):
    parsed = []
    for row in rows:
        parsed.append({field: row.get(source, "") if source else "" for field, source in mapping.items()})
    return parsed


def parse_import_contact_date(row):
    day = clean_text(row.get("dia"))
    month = clean_text(row.get("mes"))
    year = clean_text(row.get("ano"))
    if not (day and month and year):
        return None
    try:
        return date(int(float(year)), int(float(month)), int(float(day)))
    except (TypeError, ValueError):
        return None


def normalize_lead(row):
    cidade_raw = clean_text(row.get("cidade") or row.get("localidade"))
    cidade = normalize_city(extract_main_city(cidade_raw) or cidade_raw) if cidade_raw else "Sem cidade"
    nome_cliente = clean_text(row.get("nome_cliente"))
    empresa = clean_text(row.get("nome_empresa"))
    nome_empresa = empresa or nome_cliente or "Sem nome"
    tipo_cliente = normalize_client_type(row.get("tipo_cliente") or row.get("categoria") or "Outro")
    latitude = parse_float_optional(row.get("latitude"))
    longitude = parse_float_optional(row.get("longitude"))
    observacoes = clean_text(row.get("observacoes"))
    observacoes_contacto = clean_text(row.get("observacoes_contacto"))
    estado_sugerido, classificacao, motivo = classify_observations(observacoes, observacoes_contacto)
    lead = {
        "nome_cliente": nome_cliente or nome_empresa,
        "area_negocio": tipo_cliente,
        "cidade": cidade,
        "empresa": empresa,
        "nome_empresa": nome_empresa,
        "tipo_cliente": tipo_cliente,
        "morada": clean_text(row.get("morada")),
        "codigo_postal": clean_text(row.get("codigo_postal")),
        "localidade": cidade,
        "contacto": nome_cliente or nome_empresa,
        "telefone": normalize_phone_list(row.get("telefone")),
        "email": clean_text(row.get("email")).lower(),
        "comercial_responsavel": clean_text(row.get("comercial_responsavel")) or UNASSIGNED_COMMERCIAL,
        "categoria": clean_text(row.get("categoria")),
        "nif": clean_text(row.get("nif")),
        "observacoes": observacoes,
        "observacoes_contacto": observacoes_contacto,
        "reuniao_info": clean_text(row.get("reuniao_info")),
        "latitude": latitude,
        "longitude": longitude,
        "estado": estado_sugerido,
        "estado_lead": estado_sugerido,
        "classificacao_observacao": classificacao,
        "motivo_classificacao": motivo,
        "_data_ultimo_contacto": parse_import_contact_date(row),
    }
    if estado_sugerido == "Adiar contacto":
        lead["data_novo_contacto"] = date.today() + timedelta(days=30)
    return lead


def build_duplicate_indexes(exclude_lead_id=None):
    phone_index = {}
    fallback_index = {}
    for lead in Lead.query.all():
        if exclude_lead_id and lead.id == exclude_lead_id:
            continue
        for phone in extract_phone_numbers(lead.telefone):
            phone_index.setdefault(phone, lead)
        key = dedupe_fallback_key(lead.to_dict())
        if key:
            fallback_index.setdefault(key, lead)
    return phone_index, fallback_index


def detect_duplicate(lead, phone_index=None, fallback_index=None, exclude_lead_id=None):
    phone_index = phone_index or {}
    fallback_index = fallback_index or {}
    for phone in extract_phone_numbers(lead.get("telefone")):
        duplicate = phone_index.get(phone)
        if duplicate and duplicate.id != exclude_lead_id:
            return duplicate, f"telefone:{phone}"
    email = clean_text(lead.get("email")).lower()
    if email:
        duplicate = Lead.query.filter(db.func.lower(Lead.email) == email)
        if exclude_lead_id:
            duplicate = duplicate.filter(Lead.id != exclude_lead_id)
        duplicate = duplicate.first()
        if duplicate:
            return duplicate, f"email:{email}"
    nif = normalize_phone(lead.get("nif"))
    if nif:
        for existing in Lead.query.all():
            if existing.id == exclude_lead_id:
                continue
            if normalize_phone(existing.nif) == nif:
                return existing, f"nif:{nif}"
    key = dedupe_fallback_key(lead)
    duplicate = fallback_index.get(key) if key else None
    if duplicate and duplicate.id != exclude_lead_id:
        return duplicate, key
    return None, ""


def update_manual_lead(lead, lead_data):
    editable_fields = {
        "nome_cliente",
        "area_negocio",
        "cidade",
        "empresa",
        "nome_empresa",
        "tipo_cliente",
        "morada",
        "codigo_postal",
        "localidade",
        "contacto",
        "telefone",
        "email",
        "nif",
        "observacoes",
        "observacoes_contacto",
        "estado",
        "estado_lead",
        "prioridade",
        "tags",
    }
    for field in editable_fields:
        setattr(lead, field, lead_data[field])


def index_saved_lead(lead_obj, phone_index, fallback_index):
    for phone in extract_phone_numbers(lead_obj.telefone):
        phone_index.setdefault(phone, lead_obj)
    key = dedupe_fallback_key(lead_obj.to_dict())
    if key:
        fallback_index.setdefault(key, lead_obj)


def save_or_update_lead(lead, phone_index=None, fallback_index=None):
    duplicate, reason = detect_duplicate(lead, phone_index, fallback_index)
    imported_contact_date = lead.pop("_data_ultimo_contacto", None)
    if duplicate:
        updated = update_empty_fields(duplicate, lead)
        add_history(
            duplicate,
            "Duplicado encontrado",
            f"Lead duplicada encontrada na importação; dados atualizados. Motivo: {reason}.",
        )
        if updated:
            add_history(duplicate, "Lead atualizada por duplicado", "Campos vazios foram preenchidos a partir da importacao.")
        if imported_contact_date:
            add_history(
                duplicate,
                "Contacto registado em importacao",
                "Data reconstruida a partir das colunas DIA/MES/ANO.",
                created_at=datetime.combine(imported_contact_date, datetime.min.time()),
                tipo_acao="import_excel",
            )
        return duplicate, False, updated, reason

    lead_obj = Lead(**lead)
    db.session.add(lead_obj)
    db.session.flush()
    add_history(lead_obj, "Lead importada", "Lead criada por importacao de ficheiro.")
    if lead.get("classificacao_observacao"):
        add_history(lead_obj, "Classificacao automatica", f"{lead.get('classificacao_observacao')}. {lead.get('motivo_classificacao')}")
    if imported_contact_date:
        add_history(
            lead_obj,
            "Contacto registado em importacao",
            "Data reconstruida a partir das colunas DIA/MES/ANO.",
            created_at=datetime.combine(imported_contact_date, datetime.min.time()),
            tipo_acao="import_excel",
        )
    if phone_index is not None and fallback_index is not None:
        index_saved_lead(lead_obj, phone_index, fallback_index)
    return lead_obj, True, False, ""


def safe_cache_geocoding(query, latitude, longitude):
    try:
        with db.session.no_autoflush:
            cached = GeocodingCache.query.filter_by(query_text=query).first()
        if cached:
            if not valid_coordinates(cached.latitude, cached.longitude) and valid_coordinates(latitude, longitude):
                cached.latitude = latitude
                cached.longitude = longitude
            return cached
        cached = GeocodingCache(query_text=query, latitude=latitude, longitude=longitude)
        db.session.add(cached)
        db.session.flush()
        return cached
    except IntegrityError:
        db.session.rollback()
        with db.session.no_autoflush:
            cached = GeocodingCache.query.filter_by(query_text=query).first()
        if cached and not valid_coordinates(cached.latitude, cached.longitude) and valid_coordinates(latitude, longitude):
            cached.latitude = latitude
            cached.longitude = longitude
            db.session.flush()
        return cached
    except Exception as exc:
        db.session.rollback()
        print(f"[IMPORT] cache geocoding ignorada para '{query}': {exc}", flush=True)
        return None


def geocode_candidates(location):
    if isinstance(location, dict):
        raw_city = clean_text(location.get("raw") or location.get("city"))
        city = clean_text(location.get("city"))
        postal = clean_text(location.get("postal"))
    else:
        raw_city = clean_text(location)
        city = clean_text(location)
        postal = ""

    pieces = []
    sequence = LOCALITY_FALLBACK_SEQUENCES.get(normalize_lookup(raw_city), [])
    pieces.extend(sequence)
    for value in [raw_city, city, normalize_city(raw_city), extract_main_city(raw_city), extract_main_city(city)]:
        value = clean_text(value)
        if value:
            pieces.append(value)

    for segment in re.split(r"\s*[-–—/|]\s*", raw_city):
        segment = clean_text(segment)
        if segment:
            pieces.extend([segment, normalize_city(segment)])

    hint = geo_hint_from_postal(postal)
    if hint:
        pieces.append(hint)

    candidates = []
    seen = set()
    for item in pieces:
        item = clean_text(item.replace("-", " "))
        key = normalize_lookup(item)
        if not item or key in INVALID_CITY_VALUES:
            continue
        if key in GEO_CITY_MANUAL:
            item = GEO_CITY_MANUAL[key]
            key = normalize_lookup(item)
        if key not in seen:
            candidates.append(item)
            seen.add(key)
    return candidates


def simple_geocode_city(location, import_cache):
    candidates = geocode_candidates(location)
    for city in candidates:
        key = normalize_lookup(city)
        fallback = region_fallback_lookup(key)
        query = regional_cache_query(key) if fallback else f"{city}, Portugal"
        if query in import_cache:
            cached_result = import_cache[query]
            if cached_result:
                return cached_result
            continue

        try:
            with db.session.no_autoflush:
                cached = GeocodingCache.query.filter_by(query_text=query).first()
            if cached and valid_coordinates(cached.latitude, cached.longitude):
                result_method = "fallback" if query.startswith("regional:") else "cache"
                result = {"latitude": cached.latitude, "longitude": cached.longitude, "query": query, "cache": True, "method": result_method, "resolved_city": city}
                import_cache[query] = result
                return result

            if fallback:
                safe_cache_geocoding(query, fallback["lat"], fallback["lon"])
                result = {"latitude": fallback["lat"], "longitude": fallback["lon"], "query": query, "cache": False, "method": "fallback", "resolved_city": city}
                import_cache[query] = result
                return result

            response = requests.get(
                NOMINATIM_URL,
                params={"q": query, "format": "json", "limit": 1, "countrycodes": "pt"},
                headers={"User-Agent": NOMINATIM_USER_AGENT},
                timeout=5,
            )
            response.raise_for_status()
            payload = response.json()
            if payload:
                lat = parse_float_optional(payload[0].get("lat"))
                lon = parse_float_optional(payload[0].get("lon"))
                if valid_coordinates(lat, lon):
                    safe_cache_geocoding(query, lat, lon)
                    result = {"latitude": lat, "longitude": lon, "query": query, "cache": False, "method": "geocoding", "resolved_city": city}
                    import_cache[query] = result
                    return result
            safe_cache_geocoding(query, None, None)
            import_cache[query] = None
        except Exception as exc:
            db.session.rollback()
            import_cache[query] = None
            print(f"[IMPORT] geocoding falhou para {city}: {exc}", flush=True)
            continue
    return None


def geocode_unique_cities(cities, limit=None, progress=None):
    unique = []
    seen = set()
    for item in cities:
        raw = item.get("raw") if isinstance(item, dict) else item
        city = item.get("city") if isinstance(item, dict) else item
        postal = item.get("postal") if isinstance(item, dict) else ""
        normalized = normalize_city(extract_main_city(raw) or city or raw)
        if (not normalized or normalize_lookup(normalized) in INVALID_CITY_VALUES) and postal:
            normalized = geo_hint_from_postal(postal)
        key = normalize_lookup(normalized)
        if normalized and key not in INVALID_CITY_VALUES and key not in seen:
            unique.append({"city": normalized, "raw": raw, "postal": postal})
            seen.add(key)
    selected = unique if limit is None else unique[:limit]
    print(f"[IMPORT] geocoding cidades unicas: {len(selected)}/{len(unique)}", flush=True)
    results = {}
    cache = {}
    for index, city in enumerate(selected, start=1):
        label = city["city"]
        if progress:
            progress(
                phase=f"Geocodificar {label}...",
                percent=72 + round((index - 1) / max(len(selected), 1) * 24),
                cities_total=len(selected),
                cities_geocoded=index - 1,
                current_city=label,
            )
        result = simple_geocode_city(city, cache)
        if result:
            results[normalize_lookup(label)] = result
        if progress:
            progress(cities_geocoded=index)
    if progress and selected:
        progress(phase="Geocoding concluido", percent=96, current_city="")
    skipped = [] if limit is None else [item["city"] for item in unique[limit:]]
    return results, skipped


def import_summary(rows, auto_geocode=True, geocode_city_limit=None, progress=None):
    summary = {
        "read": len(rows),
        "imported": 0,
        "duplicates": 0,
        "updated": 0,
        "errors": [],
        "without_coordinates": 0,
        "with_coordinates": 0,
        "geocoded": 0,
        "geocoding_errors": [],
        "unique_cities": 0,
        "cities_geocoded": [],
        "cities_fallback": [],
        "cities_not_found": [],
        "city_normalization": [],
        "classification_counts": {
            "Ligar de volta": 0,
            "Adiar contacto": 0,
            "Sem interesse definitivo": 0,
            "Reunião marcada": 0,
            "Rever manualmente": 0,
        },
        "sheets_imported": [],
    }
    normalized_leads = []
    if progress:
        progress(phase="Validar linhas...", percent=18, total_rows=len(rows), processed_rows=0)
    for index, row in enumerate(rows, start=2):
        try:
            lead = normalize_lead(row)
            normalized_leads.append((index, lead, {
                "city": lead.get("cidade") or lead.get("localidade"),
                "raw": clean_text(row.get("cidade") or row.get("localidade")),
                "postal": clean_text(row.get("codigo_postal")),
            }))
            classification = lead.get("classificacao_observacao")
            if classification in summary["classification_counts"]:
                summary["classification_counts"][classification] += 1
        except Exception as exc:
            summary["errors"].append({"line": index, "error": str(exc)})
        if progress and (index == 2 or index % 25 == 0 or index - 1 == len(rows)):
            processed = min(index - 1, len(rows))
            progress(
                phase="Validar linhas...",
                percent=18 + round(processed / max(len(rows), 1) * 17),
                processed_rows=processed,
                errors=len(summary["errors"]),
            )

    phone_index, fallback_index = build_duplicate_indexes()
    saved_leads = []
    if progress:
        progress(phase="Verificar duplicados...", percent=38, processed_rows=0)
    for index, lead, location in normalized_leads:
        try:
            lead_obj, created, updated, reason = save_or_update_lead(lead, phone_index, fallback_index)
            saved_leads.append((lead_obj.id, location))
            if created:
                summary["imported"] += 1
                print(f"[IMPORT] nova lead criada id={lead_obj.id} telefone={lead_obj.telefone or '-'}", flush=True)
            else:
                summary["duplicates"] += 1
                if updated:
                    summary["updated"] += 1
                    index_saved_lead(lead_obj, phone_index, fallback_index)
                print(f"[IMPORT] duplicado id={lead_obj.id} motivo={reason}", flush=True)
            db.session.commit()
            if progress and (index == 2 or index % 10 == 0 or index - 1 == len(normalized_leads)):
                processed = min(index - 1, len(normalized_leads))
                progress(
                    phase="Guardar leads...",
                    percent=38 + round(processed / max(len(normalized_leads), 1) * 32),
                    processed_rows=processed,
                    imported=summary["imported"],
                    updated=summary["updated"],
                    duplicates=summary["duplicates"],
                    errors=len(summary["errors"]),
                )
        except Exception as exc:
            db.session.rollback()
            phone_index, fallback_index = build_duplicate_indexes()
            summary["errors"].append({"line": index, "error": str(exc)})
            print(f"[IMPORT] erro linha {index}: {exc}", flush=True)
            continue

    cities = []
    for lead_id, location in saved_leads:
        lead_obj = db.session.get(Lead, lead_id)
        if lead_obj and not valid_coordinates(lead_obj.latitude, lead_obj.longitude):
            cities.append(location)

    summary["unique_cities"] = len({
        normalize_lookup(
            normalize_city(extract_main_city(item.get("raw")) or item.get("city") or item.get("raw"))
            or geo_hint_from_postal(item.get("postal"))
        )
        for item in cities
        if item
    })

    geocoded = {}
    skipped = []
    if auto_geocode and cities:
        if progress:
            progress(phase="Geocodificar cidades...", percent=72, cities_total=summary["unique_cities"])
        geocoded, skipped = geocode_unique_cities(cities, geocode_city_limit, progress=progress)
        summary["cities_geocoded"] = sorted({
            item["city"]
            for item in cities
            if normalize_lookup(item["city"]) in geocoded
            and geocoded[normalize_lookup(item["city"])].get("method") != "fallback"
        })
        summary["cities_fallback"] = sorted({
            item["city"]
            for item in cities
            if normalize_lookup(item["city"]) in geocoded
            and geocoded[normalize_lookup(item["city"])].get("method") == "fallback"
        })
        summary["cities_not_found"] = sorted(set(skipped))
    elif progress:
        progress(phase="Geocoding ignorado", percent=95)

    cities_without_coordinates = set(summary["cities_not_found"])
    for lead_id, location in saved_leads:
        lead_obj = db.session.get(Lead, lead_id)
        if not lead_obj:
            continue
        city_key = normalize_lookup(lead_obj.cidade)
        postal_key = normalize_lookup(geo_hint_from_postal(lead_obj.codigo_postal))
        geocode_key = city_key if city_key in geocoded else postal_key
        if not valid_coordinates(lead_obj.latitude, lead_obj.longitude) and geocode_key in geocoded:
            lead_obj.latitude = geocoded[geocode_key]["latitude"]
            lead_obj.longitude = geocoded[geocode_key]["longitude"]
            summary["geocoded"] += 1
        if valid_coordinates(lead_obj.latitude, lead_obj.longitude):
            summary["with_coordinates"] += 1
        else:
            summary["without_coordinates"] += 1
            if lead_obj.cidade:
                cities_without_coordinates.add(lead_obj.cidade)

    summary["cities_not_found"] = sorted(cities_without_coordinates)
    db.session.commit()
    if progress:
        progress(
            phase="Resumo final...",
            percent=99,
            processed_rows=len(rows),
            imported=summary["imported"],
            updated=summary["updated"],
            duplicates=summary["duplicates"],
            errors=len(summary["errors"]),
        )
    return summary


def validate_and_import_rows(rows, auto_geocode=True, geocode_city_limit=None):
    return import_summary(rows, auto_geocode=auto_geocode, geocode_city_limit=geocode_city_limit)


def safe_perf_log(message):
    try:
        logger.info("%s", clean_text(message))
    except (OSError, ValueError):
        pass


def safe_options_query(name, fallback, factory):
    try:
        value = factory()
        if value is None:
            return fallback
        return value
    except Exception as exc:
        try:
            db.session.rollback()
        except Exception:
            pass
        safe_perf_log(f"[OPTIONS] {name} fallback: {exc.__class__.__name__}")
        return fallback


def distinct_non_empty(*columns):
    values = set()
    for column in columns:
        try:
            rows = db.session.query(column).filter(column.isnot(None), column != "").distinct().all()
        except Exception as exc:
            try:
                db.session.rollback()
            except Exception:
                pass
            safe_perf_log(f"[OPTIONS] distinct fallback: {exc.__class__.__name__}")
            continue
        values.update(clean_text(row[0]) for row in rows if clean_text(row[0]))
    return sorted(values)


def split_distinct_tags(column):
    tags = set()
    try:
        rows = db.session.query(column).filter(column.isnot(None), column != "").distinct().all()
    except Exception as exc:
        try:
            db.session.rollback()
        except Exception:
            pass
        safe_perf_log(f"[OPTIONS] tags fallback: {exc.__class__.__name__}")
        return []
    for (value,) in rows:
        tags.update(tag.strip() for tag in (value or "").split(",") if tag.strip())
    return sorted(tags)


def get_options():
    start = time.perf_counter()
    existing_commercials = safe_options_query(
        "comerciais",
        [],
        lambda: [item for item in distinct_non_empty(Lead.comercial_responsavel) if item != "Outro"],
    )
    options = {
        "localidades": safe_options_query("localidades", [], lambda: distinct_non_empty(Lead.cidade, Lead.localidade)),
        "tipos": safe_options_query("tipos", [], lambda: distinct_non_empty(Lead.area_negocio, Lead.tipo_cliente)),
        "estados": safe_options_query("estados", LEAD_STATES, lambda: distinct_non_empty(Lead.estado) or LEAD_STATES),
        "classificacoes": CLASSIFICATION_FILTERS,
        "comerciais": ["Todos", UNASSIGNED_COMMERCIAL] + existing_commercials,
        "mapa_comerciais": MAP_COMMERCIALS,
        "tags": safe_options_query("tags", sorted(TAG_OPTIONS), lambda: sorted(set(TAG_OPTIONS) | set(split_distinct_tags(Lead.tags)))),
        "insight_tags": safe_options_query(
            "insight_tags",
            sorted(INSIGHT_TAG_OPTIONS),
            lambda: sorted(set(INSIGHT_TAG_OPTIONS) | set(split_distinct_tags(Lead.insight_tags))),
        ),
    }
    safe_perf_log(f"[PERF] get_options {time.perf_counter() - start:.3f}s")
    return options


def minimal_map_metrics():
    start = time.perf_counter()
    total = db.session.query(db.func.count(Lead.id)).scalar() or 0
    safe_perf_log(f"[PERF] minimal_map_metrics {time.perf_counter() - start:.3f}s")
    return {
        "total": total,
        "active": 0,
        "meetings": 0,
        "crm": 0,
        "postponed_future": 0,
        "without_commercial": 0,
        "forgotten": 0,
        "avg_nearby": 0,
        "plans": 0,
    }


def minimal_metrics():
    leads = Lead.query.all()
    forgotten = forgotten_leads()
    active_with_coords = [lead for lead in leads if is_active_lead(lead) and lead.latitude is not None and lead.longitude is not None]
    nearby_counts = [len(nearby_active_leads(lead, 10)) for lead in active_with_coords]
    return {
        "total": len(leads),
        "active": sum(1 for lead in leads if is_active_lead(lead)),
        "meetings": sum(1 for lead in leads if lead.estado == "Reunião marcada"),
        "crm": sum(1 for lead in leads if normalize_legacy_state(lead.estado) == "Já tratado / no CRM"),
        "postponed_future": sum(1 for lead in leads if lead.estado == "Adiar contacto" and lead.data_novo_contacto and lead.data_novo_contacto > date.today()),
        "without_commercial": sum(1 for lead in leads if not lead.comercial_responsavel or lead.comercial_responsavel == "Outro"),
        "forgotten": len(forgotten),
        "avg_nearby": round(sum(nearby_counts) / len(nearby_counts), 1) if nearby_counts else 0,
        "plans": PlanoReunioes.query.count(),
    }


def history_matches(entry, tokens):
    text = normalize_lookup(" ".join([
        entry.tipo_acao or "",
        entry.acao or "",
        entry.resultado or "",
        entry.observacao or "",
    ]))
    return any(token in text for token in tokens)


def admin_overview_context():
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    since = datetime.utcnow() - timedelta(days=30)
    leads_snapshot = Lead.query.all()
    total_leads = len(leads_snapshot)
    active_leads = sum(1 for lead in leads_snapshot if is_active_lead(lead, today))
    inactive_leads = total_leads - active_leads
    contacted_leads = Lead.query.filter(Lead.estado != "Por contactar").count()
    scheduled_leads = Lead.query.filter(
        (Lead.data_novo_contacto.isnot(None)) | (Lead.data_reuniao.isnot(None))
    ).count()
    pending_followups = Lead.query.filter(
        Lead.data_novo_contacto.isnot(None),
        Lead.data_novo_contacto <= today,
    ).count()
    active_users_query = User.query.filter(User.ativo.is_(True))
    active_users = active_users_query.count()
    today_activity = HistoricoLead.query.filter(HistoricoLead.created_at >= today_start).all()
    treated_today_tokens = ["reuniao_marcada", "reuniao", "crm", "cliente existente", "sem interesse", "estado alterado"]
    treated_today = sum(1 for entry in today_activity if history_matches(entry, treated_today_tokens))
    recent_activity = (
        HistoricoLead.query.options(joinedload(HistoricoLead.user), joinedload(HistoricoLead.lead))
        .order_by(HistoricoLead.created_at.desc())
        .limit(8)
        .all()
    )
    operators = User.query.filter_by(role="comercial", ativo=True).order_by(User.nome.asc()).all()
    operator_ids = [user.id for user in operators]
    performance_entries = (
        HistoricoLead.query.filter(
            HistoricoLead.user_id.in_(operator_ids) if operator_ids else db.text("0=1"),
            HistoricoLead.created_at >= since,
        )
        .order_by(HistoricoLead.created_at.desc())
        .all()
    )
    entries_by_user = {}
    for entry in performance_entries:
        entries_by_user.setdefault(entry.user_id, []).append(entry)

    contact_tokens = ["contactada", "contactado", "contacto", "chamada", "telefonema", "email", "whatsapp", "sem resposta", "nao atendeu"]
    meeting_tokens = ["reuniao_marcada", "reuniao marcada", "agendada", "marcada reuniao", "reuniao", "crm registado"]
    followup_tokens = ["followup", "follow-up", "reagendado", "reagendada", "novo contacto", "ligar mais tarde", "adiar contacto"]
    operator_performance = []
    for user in operators:
        entries = entries_by_user.get(user.id, [])
        contacts = sum(1 for entry in entries if history_matches(entry, contact_tokens))
        meetings = sum(1 for entry in entries if history_matches(entry, meeting_tokens))
        followups = sum(1 for entry in entries if history_matches(entry, followup_tokens))
        conversion_rate = round((meetings / contacts) * 100) if contacts else 0
        operator_performance.append({
            "user": user,
            "total_acoes": len(entries),
            "contactos": contacts,
            "reunioes": meetings,
            "followups": followups,
            "taxa_concretizacao": conversion_rate,
            "ultima_atividade": entries[0].created_at if entries else None,
        })
    operator_performance.sort(key=lambda item: (item["total_acoes"], item["reunioes"], item["contactos"]), reverse=True)
    return {
        "metrics": {
            "total_leads": total_leads,
            "active_leads": active_leads,
            "inactive_leads": inactive_leads,
            "contacted_leads": contacted_leads,
            "scheduled_leads": scheduled_leads,
            "pending_followups": pending_followups,
            "treated_today": treated_today,
            "active_users": active_users,
        },
        "users_by_role": {
            "admin": active_users_query.filter(User.role == "admin").count(),
            "comercial": active_users_query.filter(User.role == "comercial").count(),
        },
        "operator_performance": operator_performance,
        "recent_users": active_users_query.order_by(User.created_at.desc()).limit(5).all(),
        "recent_activity": recent_activity,
    }


def personal_performance_context(user):
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    since_7 = today_start - timedelta(days=6)
    contact_tokens = ["contactada", "contactado", "contacto", "chamada", "telefonema", "email", "whatsapp", "sem resposta", "nao atendeu"]
    meeting_tokens = ["reuniao_marcada", "reuniao marcada", "agendada", "marcada reuniao", "reuniao", "crm registado"]

    recent_window = (
        HistoricoLead.query.options(joinedload(HistoricoLead.lead))
        .filter(
            HistoricoLead.user_id == user.id,
            HistoricoLead.created_at >= since_7,
        )
        .order_by(HistoricoLead.created_at.desc())
        .limit(250)
        .all()
    )
    timeline = (
        HistoricoLead.query.options(joinedload(HistoricoLead.lead))
        .filter(HistoricoLead.user_id == user.id)
        .order_by(HistoricoLead.created_at.desc())
        .limit(12)
        .all()
    )
    contacts_today = sum(
        1
        for entry in recent_window
        if entry.created_at and entry.created_at >= today_start and history_matches(entry, contact_tokens)
    )
    contacts_7 = sum(1 for entry in recent_window if history_matches(entry, contact_tokens))
    meetings_7 = sum(1 for entry in recent_window if history_matches(entry, meeting_tokens))
    conversion_rate = round((meetings_7 / contacts_7) * 100) if contacts_7 else 0
    pending_followups = (
        db.session.query(Lead.id)
        .join(HistoricoLead, HistoricoLead.lead_id == Lead.id)
        .filter(
            HistoricoLead.user_id == user.id,
            Lead.data_novo_contacto.isnot(None),
            Lead.data_novo_contacto <= today,
            or_(HistoricoLead.tipo_acao == "followup_reagendado", HistoricoLead.acao.ilike("%Adiar%")),
        )
        .distinct()
        .count()
    )

    return {
        "metrics": {
            "contacts_today": contacts_today,
            "contacts_7": contacts_7,
            "meetings": meetings_7,
            "conversion_rate": conversion_rate,
            "pending_followups": pending_followups,
            "last_activity": timeline[0].created_at if timeline else None,
        },
        "timeline": timeline,
        "has_history": bool(timeline),
    }


def invalid_location_value(value):
    key = normalize_lookup(value)
    return not clean_text(value) or key in INVALID_CITY_VALUES or key == normalize_lookup("Sem cidade")


def needs_coordinate_review(lead):
    if valid_coordinates(lead.latitude, lead.longitude):
        return False
    return (
        invalid_location_value(lead.cidade or lead.localidade)
        or "Ignorar mapa" in lead.tag_list()
        or not valid_coordinates(lead.latitude, lead.longitude)
    )


def coordinate_review_reason(lead):
    if "Ignorar mapa" in lead.tag_list():
        return "ignorar no mapa"
    if invalid_location_value(lead.cidade or lead.localidade):
        return "cidade em falta"
    if not clean_text(lead.cidade or lead.localidade or lead.morada or lead.codigo_postal):
        return "sem dados suficientes"
    if not valid_coordinates(lead.latitude, lead.longitude):
        if lead.historico and any("geocoding" in normalize_lookup(item.acao) and "erro" in normalize_lookup(item.acao) for item in lead.historico):
            return "geocoding falhou"
        return "cidade não encontrada"
    return ""


def geocode_single_lead(lead):
    if valid_coordinates(lead.latitude, lead.longitude):
        return True, "Lead ja tem coordenadas."
    result = simple_geocode_city({
        "city": lead.cidade or lead.localidade,
        "raw": lead.cidade or lead.localidade,
        "postal": lead.codigo_postal,
    }, {})
    if result:
        lead.latitude = result["latitude"]
        lead.longitude = result["longitude"]
        add_history(lead, "Geocoding manual", f"Query: {result.get('query', '')}. Resultado: {lead.latitude}, {lead.longitude}.")
        return True, "Coordenadas atualizadas."
    add_history(lead, "Erro de geocoding manual", f"Nao foi possivel geocodificar {lead.cidade or lead.localidade or lead.codigo_postal or 'sem localizacao'}.")
    return False, "Geocoding falhou."


def normalize_legacy_state(state):
    value = clean_text(state)
    mapping = {
        "Contactado": "Ligar de volta",
        "Reunião marcada": "Já tratado / no CRM",
        "ReuniÃ£o marcada": "Já tratado / no CRM",
        "Cliente existente": "Já tratado / no CRM",
        "Sem interesse definitivo": "Sem interesse",
    }
    return mapping.get(value, value)


def normalize_commercial_key(value):
    text = normalize_lookup(value)
    text = re.sub(r"[^a-z0-9]+", "", text)
    aliases = {
        "": "sem_comercial",
        "outro": "sem_comercial",
        normalize_lookup(UNASSIGNED_COMMERCIAL).replace(" ", ""): "sem_comercial",
        "semcomercial": "sem_comercial",
        "semcomercialatribuido": "sem_comercial",
        "semcomercialatribudo": "sem_comercial",
        "ines": "ines",
        "ins": "ines",
        "bruno": "bruno",
        "flavia": "flavia",
        "flvia": "flavia",
        "miriam": "miriam",
        "setil": "setil",
    }
    return aliases.get(text, text)


def display_commercial(value):
    key = normalize_commercial_key(value)
    labels = {item["value"]: item["label"] for item in MAP_COMMERCIALS}
    return labels.get(key, clean_text(value) or "Sem comercial")


def commercial_label_from_key(key):
    labels = {item["value"]: item["label"] for item in MAP_COMMERCIALS}
    return labels.get(key, UNASSIGNED_COMMERCIAL)


def build_manual_lead(form):
    cidade_raw = clean_text(form.get("cidade"))
    cidade = normalize_city(extract_main_city(cidade_raw) or cidade_raw) if cidade_raw else ""
    nome_cliente = clean_text(form.get("nome_cliente"))
    empresa = clean_text(form.get("empresa"))
    nome_empresa = empresa or nome_cliente
    tipo_cliente = normalize_client_type(form.get("area_negocio") or "Outro")
    estado = clean_text(form.get("estado")) or "Por contactar"
    if estado not in LEAD_STATES:
        estado = "Por contactar"
    return {
        "nome_cliente": nome_cliente or nome_empresa,
        "area_negocio": tipo_cliente,
        "cidade": cidade or "Sem cidade",
        "empresa": empresa,
        "nome_empresa": nome_empresa,
        "tipo_cliente": tipo_cliente,
        "morada": clean_text(form.get("morada")),
        "codigo_postal": clean_text(form.get("codigo_postal")),
        "localidade": cidade or "Sem cidade",
        "contacto": nome_cliente or nome_empresa,
        "telefone": normalize_phone_list(form.get("telefone")),
        "email": clean_text(form.get("email")).lower(),
        "categoria": "",
        "nif": clean_text(form.get("nif")),
        "observacoes": clean_text(form.get("observacoes")),
        "observacoes_contacto": clean_text(form.get("observacoes_contacto")),
        "reuniao_info": "",
        "comercial_responsavel": UNASSIGNED_COMMERCIAL,
        "latitude": None,
        "longitude": None,
        "estado": estado,
        "estado_lead": estado,
        "prioridade": clean_text(form.get("prioridade")) or "Baixa",
        "tags": clean_text(form.get("tags")),
        "classificacao_observacao": "",
        "motivo_classificacao": "",
    }


def distance_km(a, b):
    if a.latitude is None or a.longitude is None or b.latitude is None or b.longitude is None:
        return None
    earth = 6371
    dlat = math.radians(b.latitude - a.latitude)
    dlon = math.radians(b.longitude - a.longitude)
    lat1 = math.radians(a.latitude)
    lat2 = math.radians(b.latitude)
    value = math.sin(dlat / 2) ** 2 + math.sin(dlon / 2) ** 2 * math.cos(lat1) * math.cos(lat2)
    return earth * 2 * math.atan2(math.sqrt(value), math.sqrt(1 - value))


def nearby_active_leads(base, radius):
    rows = []
    for lead in Lead.query.all():
        if lead.id == base.id or not is_active_lead(lead):
            continue
        distance = distance_km(base, lead)
        if distance is not None and distance <= radius:
            rows.append((lead, round(distance, 1)))
    return sorted(rows, key=lambda item: item[1])


def forgotten_leads(days=60):
    cutoff = datetime.utcnow() - timedelta(days=days)
    rows = []
    for lead in Lead.query.all():
        if not is_active_lead(lead):
            continue
        latest = lead.historico[0].created_at if lead.historico else None
        baseline = latest or lead.created_at or datetime.utcnow()
        if latest is None or latest < cutoff:
            rows.append((lead, latest, (datetime.utcnow() - baseline).days))
    return rows


def dashboard_context():
    leads = Lead.query.order_by(Lead.updated_at.desc()).all()
    no_coords = [lead for lead in leads if needs_coordinate_review(lead)]
    active = [lead for lead in leads if is_active_lead(lead)]
    scheduled = scheduled_leads_context(limit_today=5)
    zone_counts = Counter((lead.cidade or lead.localidade or "Sem cidade") for lead in active if valid_coordinates(lead.latitude, lead.longitude))
    best_zone = zone_counts.most_common(1)[0] if zone_counts else ("Sem zona ativa", 0)
    today = date.today()
    today_history = HistoricoLead.query.filter(db.func.date(HistoricoLead.created_at) == today.isoformat()).all()
    commercial_counts = Counter(display_commercial(lead.comercial_responsavel) for lead in active)
    recent_imports = HistoricoLead.query.filter(HistoricoLead.acao == "Lead importada").order_by(HistoricoLead.created_at.desc()).limit(5).all()
    inbox_new = [
        lead for lead in active
        if not lead.historico and normalize_legacy_state(lead.estado) == "Por contactar"
    ][:8]
    unassigned = [lead for lead in active if lead_has_no_commercial(lead)][:8]
    no_coords_priority = sorted(
        [lead for lead in no_coords if is_active_lead(lead)],
        key=lambda lead: operational_score(lead),
        reverse=True,
    )[:8]
    week_start = datetime.utcnow() - timedelta(days=7)
    week_history = [item for item in HistoricoLead.query.filter(HistoricoLead.created_at >= week_start).all()]
    week_contacts = sum(1 for item in week_history if ("contact" in normalize_lookup(item.acao) or "ligar" in normalize_lookup(item.acao)))
    week_meetings = sum(1 for item in week_history if ("reuni" in normalize_lookup(item.acao) or "crm" in normalize_lookup(item.acao)))
    week_state_changes = sum(1 for item in week_history if ("estado" in normalize_lookup(item.acao)))
    return {
        "metrics": {
            "total": len(leads),
            "active": len(active),
            "forgotten": len(forgotten_leads(60)),
            "without_coordinates": len(no_coords),
            "without_commercial": sum(1 for lead in leads if lead_has_no_commercial(lead)),
            "contacts_today": sum(1 for item in today_history if "contact" in normalize_lookup(item.acao) or "ligar" in normalize_lookup(item.acao)),
            "meetings": sum(1 for lead in leads if "reun" in normalize_lookup(lead.estado)),
            "unanswered": sum(1 for lead in leads if "nao atendeu" in normalize_lookup(" ".join([lead.observacoes or "", lead.observacoes_contacto or ""]))),
            "followups_today": scheduled["today_count"],
            "followups_overdue": scheduled["overdue_count"],
        },
        "best_zone": {"name": best_zone[0], "count": best_zone[1]},
        "recent_leads": leads[:6],
        "recent_imports": recent_imports,
        "commercial_counts": commercial_counts.most_common(5),
        "score_preview": sorted([(lead, operational_score(lead)) for lead in leads if lead.latitude is not None or lead.longitude is not None], key=lambda item: item[1], reverse=True)[:6],
        "followups_today": scheduled["today"],
        "followups_overdue_count": scheduled["overdue_count"],
        "inbox_new": inbox_new,
        "unassigned": unassigned,
        "no_coords_priority": no_coords_priority,
        "weekly_summary": {
            "contacts": week_contacts,
            "meetings": week_meetings,
            "state_changes": week_state_changes,
            "period_days": 7,
        },
    }


def city_suggestion(value):
    original = clean_text(value)
    suggested = normalize_city(extract_main_city(original) or original)
    if suggested and normalize_lookup(suggested) != normalize_lookup(original):
        return suggested
    return ""


def coordinate_quality_context(rows):
    total = Lead.query.count()
    with_coords = Lead.query.filter(Lead.latitude.isnot(None), Lead.longitude.isnot(None)).count()
    reason_counts = Counter(item["motivo"] for item in rows)
    city_counts = Counter((item["lead"].cidade or item["lead"].localidade or "Sem cidade") for item in rows)
    suggestions = []
    for item in rows:
        lead = item["lead"]
        suggested = city_suggestion(lead.cidade or lead.localidade)
        if suggested:
            suggestions.append({"lead": lead, "suggested": suggested})
    return {
        "total": total,
        "with_coords": with_coords,
        "without_coords": len(rows),
        "geocoded_percent": round((with_coords / total) * 100, 1) if total else 100,
        "total_errors": len(rows),
        "problem_cities": city_counts.most_common(6),
        "top_failures": reason_counts.most_common(6),
        "suggestions": suggestions[:8],
    }


def search_leads(query, limit=8):
    term = normalize_lookup(query)
    if not term:
        return []
    like = f"%{clean_text(query).lower()}%"
    results = []
    scope = requested_assignment_scope()
    lead_query = apply_assignment_scope(Lead.query, scope).filter(or_(
        db.func.lower(db.func.coalesce(Lead.nome_cliente, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.nome_empresa, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.empresa, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.telefone, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.email, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.cidade, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.localidade, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.tags, "")).like(like),
        db.func.lower(db.func.coalesce(Lead.insight_tags, "")).like(like),
    ))
    for lead in lead_query.order_by(Lead.updated_at.desc(), Lead.nome_empresa.asc()).limit(limit).all():
        haystack = normalize_lookup(" ".join([
            lead.nome_cliente or "",
            lead.nome_empresa or "",
            lead.empresa or "",
            lead.telefone or "",
            lead.cidade or lead.localidade or "",
            lead.email or "",
            lead.tags or "",
            lead.insight_tags or "",
            display_commercial(lead.comercial_responsavel),
        ]))
        if term in haystack:
            results.append({
                "type": "lead",
                "id": lead.id,
                "title": lead.nome_cliente or lead.nome_empresa,
                "subtitle": f"{lead.nome_empresa or lead.empresa or '-'} · {lead.cidade or lead.localidade or '-'}",
                "phone": lead.telefone or "",
                "commercial": display_commercial(lead.comercial_responsavel),
                "url": url_for("mapa_leads", lead_id=lead.id),
            })
        if len(results) >= limit:
            break
    return results


def all_leads_payload(scope=None):
    scope = scope if scope is not None else requested_assignment_scope()
    query = apply_assignment_scope(Lead.query.options(joinedload(Lead.assigned_to), selectinload(Lead.historico)), scope)
    leads = [
        lead
        for lead in query.order_by(Lead.nome_empresa.asc()).all()
        if is_active_lead(lead)
    ]
    payload = []
    tags = set()
    states = set()
    for lead in leads:
        item = lead.to_dict(include_history=False)
        item["estado"] = normalize_legacy_state(item["estado"])
        item["ativa"] = is_active_lead(lead)
        item["agendada"] = is_scheduled_lead(lead)
        item["tem_coordenadas"] = valid_coordinates(lead.latitude, lead.longitude)
        item["comercial_responsavel"] = display_commercial(item.get("comercial_responsavel"))
        payload.append(item)
        states.add(item["estado"])
        for tag in item.get("tags", []):
            if tag:
                tags.add(tag)
    return leads, payload, {"states": sorted(states), "tags": sorted(tags)}


def all_leads_filtered_for_export(scope=None):
    _, leads, _ = all_leads_payload(scope)
    query = normalize_lookup(request.args.get("q"))
    state = request.args.get("state") or ""
    heat = request.args.get("heat") or ""
    tag = request.args.get("tag") or ""
    coords = request.args.get("coords") or ""
    sort = request.args.get("sort") or "important"

    def lead_name(lead):
        return lead.get("nome_cliente") or lead.get("nome") or lead.get("empresa") or lead.get("nome_empresa") or "Lead sem nome"

    def matches(lead):
        if query:
            haystack = normalize_lookup(" ".join([
                lead_name(lead),
                lead.get("nome_empresa") or "",
                lead.get("cidade") or "",
                lead.get("localidade") or "",
                lead.get("telefone") or "",
            ]))
            if query not in haystack:
                return False
        if state and lead.get("estado") != state:
            return False
        if heat and lead.get("heat_band") != heat:
            return False
        if tag and tag not in (lead.get("tags") or []):
            return False
        if coords == "with" and not lead.get("tem_coordenadas"):
            return False
        if coords == "without" and lead.get("tem_coordenadas"):
            return False
        return True

    def schedule_value(lead):
        value = lead.get("data_novo_contacto")
        if not value:
            return (date.max.isoformat(), "99:99")
        return (value, lead.get("hora_reuniao") or "23:59")

    def importance(lead):
        value = int(lead.get("score") or 0) * 10
        if lead.get("agendada"):
            value += 140
        if not lead.get("tem_coordenadas"):
            value += 35
        if lead.get("data_novo_contacto"):
            value += 25
        if not lead.get("comercial_responsavel"):
            value += 15
        return value

    rows = [lead for lead in leads if matches(lead)]
    sorters = {
        "important": lambda lead: (-importance(lead), normalize_lookup(lead_name(lead))),
        "score_desc": lambda lead: (-(lead.get("score") or 0), normalize_lookup(lead_name(lead))),
        "next_contact": lambda lead: (*schedule_value(lead), -(lead.get("score") or 0)),
        "no_coords": lambda lead: (1 if lead.get("tem_coordenadas") else 0, -(lead.get("score") or 0)),
        "commercial": lambda lead: normalize_lookup(lead.get("comercial_responsavel") or "Sem comercial"),
        "city": lambda lead: normalize_lookup(lead.get("cidade") or lead.get("localidade") or "Sem cidade"),
        "name": lambda lead: normalize_lookup(lead_name(lead)),
    }
    return sorted(rows, key=sorters.get(sort, sorters["important"]))


def build_leads_export_workbook(leads):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Leads"
    headers = ["Nome", "Empresa", "Cidade", "Estado", "Comercial", "Próximo contacto", "Email", "Telefone"]
    worksheet.append(headers)
    for lead in leads:
        next_contact = lead.get("data_novo_contacto") or ""
        if next_contact and lead.get("hora_reuniao"):
            next_contact = f"{next_contact} {lead.get('hora_reuniao')}"
        worksheet.append([
            lead.get("nome_cliente") or lead.get("nome_empresa") or "",
            lead.get("nome_empresa") or lead.get("empresa") or "",
            lead.get("cidade") or lead.get("localidade") or "",
            lead.get("estado") or "",
            lead.get("comercial_responsavel") or "",
            next_contact,
            lead.get("email") or "",
            lead.get("telefone") or "",
        ])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(vertical="center")
    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width, 14), 34)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def operational_score(lead):
    try:
        return int(lead.operational_score())
    except Exception:
        return 0


def lightweight_heat_score(lead):
    score = 35
    estado = normalize_lookup(lead.estado)
    if lead.telefone or lead.email:
        score += 5
    if not lead.comercial_responsavel or normalize_commercial_key(lead.comercial_responsavel) == "sem_comercial":
        score += 8
    if lead.data_novo_contacto:
        today = date.today()
        if lead.data_novo_contacto < today:
            score += 30
        elif lead.data_novo_contacto == today:
            score += 18
        else:
            score += 8
    text = normalize_lookup(" ".join(filter(None, [
        lead.estado,
        lead.observacoes_contacto,
        lead.reuniao_info,
        lead.motivo_classificacao,
    ])))
    if any(token in text for token in ["reuniao", "crm", "marcada", "agendado"]):
        score += 18
    if any(token in text for token in ["nao atendeu", "sem resposta", "ligar depois", "voltar a contactar"]):
        score += 10
    if "sem interesse" in estado:
        score -= 45
    if not lead.latitude or not lead.longitude:
        score -= 5
    return max(0, min(100, score))


def heat_band_from_score(score):
    if score >= 70:
        return "hot"
    if score >= 40:
        return "warm"
    return "cold"


def heat_label_from_band(band):
    return {
        "hot": "🔥 quente",
        "warm": "🟡 morna",
        "cold": "⚪ fria",
    }.get(band, "⚪ fria")


def lead_map_payload(lead, include_history=False):
    heat_score = lightweight_heat_score(lead)
    heat_band = heat_band_from_score(heat_score)
    item = {
        "id": lead.id,
        "nome_cliente": lead.nome_cliente or lead.nome_empresa,
        "nome_empresa": lead.nome_empresa,
        "empresa": lead.empresa or "",
        "area_negocio": lead.area_negocio or lead.tipo_cliente,
        "tipo_cliente": lead.tipo_cliente,
        "cidade": lead.cidade or lead.localidade,
        "localidade": lead.localidade,
        "telefone": lead.telefone or lead.contacto or "",
        "email": lead.email or "",
        "latitude": lead.latitude,
        "longitude": lead.longitude,
        "estado": normalize_legacy_state(lead.estado),
        "prioridade": lead.prioridade or "",
        "comercial_responsavel": display_commercial(lead.comercial_responsavel),
        "comercial_key": normalize_commercial_key(lead.comercial_responsavel),
        "data_novo_contacto": lead.data_novo_contacto.isoformat() if lead.data_novo_contacto else "",
        "data_reuniao": lead.data_reuniao.isoformat() if lead.data_reuniao else "",
        "hora_reuniao": lead.hora_reuniao or "",
        "tags": lead.tag_list(),
        "insight_tags": lead.insight_tag_list(),
        "insight_note": "",
        "observacoes": "",
        "observacoes_contacto": "",
        "score": heat_score,
        "score_band": heat_band,
        "heat_score": heat_score,
        "heat_band": heat_band,
        "heat_label": heat_label_from_band(heat_band),
        "ativa": is_active_lead(lead),
        "agendada": is_scheduled_lead(lead),
        "agenda_bucket": scheduled_bucket(lead),
        "tem_coordenadas": lead.latitude is not None and lead.longitude is not None,
    }
    if include_history:
        item["historico"] = [entry.to_dict() for entry in lead.historico]
    return item


def score_to_band(score):
    if score >= 70:
        return "high"
    if score >= 40:
        return "medium"
    return "low"


def next_best_lead(base_lead=None, commercial_filter=None):
    candidates = Lead.query.all()
    base_city = normalize_lookup((base_lead.cidade or base_lead.localidade or "")) if base_lead else ""
    base_commercial = clean_text(commercial_filter or (base_lead.comercial_responsavel if base_lead else ""))
    scored = []
    for lead in candidates:
        if base_lead and lead.id == base_lead.id:
            continue
        if not eligible_planning_lead(lead):
            continue
        lead_commercial = clean_text(lead.comercial_responsavel)
        if base_commercial and lead_commercial not in {base_commercial, "", clean_text(UNASSIGNED_COMMERCIAL), "sem comercial", "sem comercial atribuido"}:
            continue
        if base_lead and valid_coordinates(base_lead.latitude, base_lead.longitude) and valid_coordinates(lead.latitude, lead.longitude):
            distance = distance_km(base_lead, lead) or 9999
        else:
            distance = 9999
        same_zone = 1 if base_city and base_city == normalize_lookup((lead.cidade or lead.localidade or "")) else 0
        scored.append((lead, distance, same_zone, operational_score(lead)))
    scored.sort(key=lambda item: (-item[3], -item[2], item[1], clean_text(item[0].nome_empresa or item[0].nome_cliente)))
    return scored[0][0] if scored else None


def lead_ignored_on_map(lead):
    return "Ignorar mapa" in lead.tag_list()


def lead_has_no_commercial(lead):
    commercial = clean_text(lead.comercial_responsavel)
    return not commercial or commercial in {"Outro", UNASSIGNED_COMMERCIAL}


def eligible_planning_lead(lead):
    state = normalize_legacy_state(lead.estado)
    return (
        is_active_lead(lead)
        and valid_coordinates(lead.latitude, lead.longitude)
        and not lead_ignored_on_map(lead)
        and state not in {"Sem interesse", "Já tratado / no CRM"}
    )


def commercial_matches(lead, selected):
    selected = clean_text(selected) or PLANNING_COMMERCIALS[0]
    commercial = clean_text(lead.comercial_responsavel)
    return commercial == selected or lead_has_no_commercial(lead)


def planning_ownership_label(lead, selected):
    return "Sem comercial" if lead_has_no_commercial(lead) else "Do comercial"


def estimate_route_km(leads):
    ordered = [lead for lead in leads if valid_coordinates(lead.latitude, lead.longitude)]
    if len(ordered) <= 1:
        return 0
    total = 0
    for index in range(1, len(ordered)):
        total += distance_km(ordered[index - 1], ordered[index]) or 0
    return round(total, 1)


def choose_default_zone(leads):
    city_groups = {}
    for lead in leads:
        city = clean_text(lead.cidade or lead.localidade) or "Sem cidade"
        city_groups.setdefault(city, []).append(lead)
    if not city_groups:
        return ""
    return sorted(city_groups.items(), key=lambda item: len(item[1]), reverse=True)[0][0]


def generate_day_contact_plan(leads, radius_km, limit, base_lead_id=None, base_city=""):
    base = next((lead for lead in leads if lead.id == base_lead_id), None) if base_lead_id else None
    zone = clean_text(base_city)
    if base:
        zone = clean_text(base.cidade or base.localidade)
    if not zone:
        zone = choose_default_zone(leads)

    pool = leads
    if zone:
        zone_key = normalize_lookup(zone)
        pool = [lead for lead in pool if normalize_lookup(lead.cidade or lead.localidade) == zone_key]
    if base:
        measured = [
            (lead, 0 if lead.id == base.id else round(distance_km(base, lead) or 0, 1))
            for lead in pool
            if lead.id == base.id or (distance_km(base, lead) is not None and distance_km(base, lead) <= radius_km)
        ]
    else:
        measured = [(lead, 0) for lead in pool]

    measured.sort(key=lambda item: (item[1], item[0].nome_cliente or item[0].nome_empresa or ""))
    selected = measured[:limit]
    return {
        "commercial": "",
        "zone": zone or "Sem zona",
        "radius": radius_km,
        "total_available": len(measured),
        "limit": limit,
        "too_many": len(measured) > limit or len(measured) > 15,
        "estimated_km": estimate_route_km([lead for lead, _distance in selected]),
        "rows": [{"lead": lead, "distance": distance} for lead, distance in selected],
    }


def day_planning_summary(leads, plan, excluded_other_count=0):
    cities = sorted({lead.cidade or lead.localidade for lead in leads if lead.cidade or lead.localidade})
    return {
        "total_leads": len(leads),
        "cities": cities,
        "cities_count": len(cities),
        "estimated_km": plan.get("estimated_km", 0) if plan else 0,
        "suggested_count": len(plan.get("rows", [])) if plan else 0,
        "excluded_other_count": excluded_other_count,
    }


def export_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contactos Zona"
    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="ECE8DF")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="1F2933", bold=True)
    sheet.append(["Alltera - Lista de contactos da zona"])
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    if rows:
        first = rows[0]
        sheet.append(["Raio/observação", first.get("raio_km", "")])
    else:
        sheet.append(["Raio/observação", ""])
    sheet.append([])
    headers = ["nome_empresa", "tipo_cliente", "morada", "codigo_postal", "localidade", "contacto", "telefone", "email", "estado", "observacoes", "distancia_km"]
    sheet.append(headers)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A{header_row}:K{sheet.max_row}"
    for index, column in enumerate(sheet.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in column) + 3, 45)
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def latest_import_summary():
    latest = HistoricoLead.query.filter(HistoricoLead.acao.in_(["Lead importada", "Duplicado encontrado"])).order_by(HistoricoLead.created_at.desc()).first()
    if not latest:
        return None
    start = latest.created_at - timedelta(minutes=30)
    end = latest.created_at + timedelta(minutes=5)
    imported = HistoricoLead.query.filter(HistoricoLead.acao == "Lead importada", HistoricoLead.created_at >= start, HistoricoLead.created_at <= end).count()
    duplicates = HistoricoLead.query.filter(HistoricoLead.acao == "Duplicado encontrado", HistoricoLead.created_at >= start, HistoricoLead.created_at <= end).count()
    return {
        "created_at": latest.created_at,
        "imported": imported,
        "duplicates": duplicates,
    }


def create_import_job(user_id=None):
    job_id = uuid.uuid4().hex
    now = time.time()
    with IMPORT_JOBS_LOCK:
        IMPORT_JOBS[job_id] = {
            "id": job_id,
            "status": "queued",
            "phase": "Preparar ficheiro...",
            "percent": 1,
            "total_rows": 0,
            "processed_rows": 0,
            "cities_total": 0,
            "cities_geocoded": 0,
            "current_city": "",
            "imported": 0,
            "updated": 0,
            "duplicates": 0,
            "errors": 0,
            "elapsed": 0,
            "eta": "",
            "summary": None,
            "error": "",
            "user_id": user_id,
            "started_at": now,
            "finished_at": None,
        }
    return job_id


def update_import_job(job_id, **changes):
    now = time.time()
    with IMPORT_JOBS_LOCK:
        job = IMPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(changes)
        started_at = job.get("started_at") or now
        elapsed = max(0, now - started_at)
        job["elapsed"] = round(elapsed, 1)
        percent = max(float(job.get("percent") or 0), 1)
        if job.get("status") not in {"done", "error"} and percent < 100:
            remaining = max(0, (elapsed / percent) * (100 - percent))
            job["eta"] = format_duration(remaining)
        elif job.get("status") == "done":
            job["eta"] = "0s"
            job["finished_at"] = job.get("finished_at") or now


def get_import_job(job_id):
    with IMPORT_JOBS_LOCK:
        job = IMPORT_JOBS.get(job_id)
        return dict(job) if job else None


def format_duration(seconds):
    seconds = int(max(0, seconds))
    if seconds < 60:
        return f"{seconds}s"
    minutes, rest = divmod(seconds, 60)
    if minutes < 60:
        return f"{minutes}m {rest}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes}m"


def run_import_job(app, job_id, filename, file_bytes, auto_geocode, sheet_name=None, user_id=None):
    with app.app_context():
        AUDIT_CONTEXT.user_id = user_id
        update_import_job(job_id, status="running", phase="Preparar ficheiro...", percent=3)
        try:
            print("[IMPORT] INICIO IMPORTACAO", flush=True)
            print(f"[IMPORT] ficheiro recebido: {filename}", flush=True)
            file_storage = FileStorage(stream=BytesIO(file_bytes), filename=filename)
            update_import_job(job_id, phase="Ler Excel...", percent=8)
            rows = read_import_file(file_storage, sheet_name=sheet_name)
            update_import_job(job_id, phase="Mapear colunas...", percent=14, total_rows=len(rows))
            print(f"[IMPORT] linhas lidas: {len(rows)}", flush=True)
            headers = list(rows[0].keys()) if rows else []
            mapping = map_columns(headers)
            print(f"[IMPORT] colunas detetadas: {mapping}", flush=True)
            parsed_rows = parse_rows(rows, mapping)

            def progress(**changes):
                update_import_job(job_id, **changes)

            summary = import_summary(parsed_rows, auto_geocode=auto_geocode, progress=progress)
            if sheet_name:
                summary["sheets_imported"] = [sheet_name]
            update_import_job(
                job_id,
                status="done",
                phase="Importacao concluida",
                percent=100,
                processed_rows=summary.get("read", 0),
                imported=summary.get("imported", 0),
                updated=summary.get("updated", 0),
                duplicates=summary.get("duplicates", 0),
                errors=len(summary.get("errors", [])),
                summary=summary,
                eta="0s",
            )
            print(
                f"[IMPORT] FIM IMPORTACAO criadas={summary['imported']} "
                f"atualizadas={summary['updated']} duplicados={summary['duplicates']} erros={len(summary['errors'])}",
                flush=True,
            )
        except Exception as exc:
            db.session.rollback()
            update_import_job(job_id, status="error", phase="Erro na importacao", error=str(exc), percent=100, eta="0s")
            print(f"[IMPORT] erro: {exc}", flush=True)
        finally:
            AUDIT_CONTEXT.user_id = None
            db.session.remove()
            print("[IMPORT] sessao fechada", flush=True)


def register_routes(app):
    @app.context_processor
    def inject_sidebar_counts():
        if request.endpoint == "login":
            return {"scheduled_sidebar_count": 0, "operational_notifications": {"count": 0, "items": []}}
        return {
            "scheduled_sidebar_count": scheduled_badge_count(),
            "operational_notifications": operational_notifications_context(),
        }

    @app.route("/registo", methods=["GET", "POST"])
    def registo():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            nome = clean_text(request.form.get("nome"))
            email = clean_text(request.form.get("email")).lower()
            password = request.form.get("password") or ""
            confirm_password = request.form.get("confirm_password") or ""

            if not nome:
                flash("Nome obrigatorio.", "error")
            elif not email:
                flash("Email obrigatorio.", "error")
            elif len(password) < 6:
                flash("A password deve ter pelo menos 6 caracteres.", "error")
            elif password != confirm_password:
                flash("As passwords nao coincidem.", "error")
            elif User.query.filter_by(email=email).first():
                flash("Este email ja existe.", "error")
            else:
                user = User(nome=nome, email=email, role="comercial", ativo=False, approval_status="pending")
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash("Pedido de registo enviado. Aguarda aprovação do administrador.", "success")
                return redirect(url_for("login"))
        return render_template("registo.html")

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            email = clean_text(request.form.get("email")).lower()
            password = request.form.get("password") or ""
            user = User.query.filter_by(email=email).first()
            if not user or not user.check_password(password):
                flash("Email ou password invalidos.", "error")
                return render_template("login.html")
            approval_status = user.approval_status or "approved"
            if approval_status == "pending":
                flash("O teu pedido ainda aguarda aprovação do administrador.", "warning")
            elif approval_status == "rejected":
                flash("O teu pedido de acesso foi rejeitado.", "error")
            elif not user.ativo:
                flash("A tua conta está inativa. Contacta o administrador.", "error")
            elif approval_status == "approved":
                login_user(user, remember=request.form.get("remember") == "1")
                next_url = request.args.get("next")
                if is_safe_next_url(next_url):
                    return redirect(next_url)
                return redirect(url_for("admin_overview" if user.role == "admin" else "mapa_leads"))
            else:
                flash("Email ou password invalidos.", "error")
        return render_template("login.html")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        flash("Sessao terminada com seguranca.", "success")
        return redirect(url_for("login"))

    @app.route("/notas-equipa", methods=["GET", "POST"])
    @login_required
    def notas_equipa():
        if request.method == "POST":
            title = clean_text(request.form.get("titulo"))
            content = clean_text(request.form.get("conteudo"))
            if not title:
                flash("Indica um titulo para a nota.", "error")
            elif not content:
                flash("Escreve o conteudo da nota.", "error")
            else:
                note = EquipaNota(titulo=title, conteudo=content, autor_id=current_user.id)
                db.session.add(note)
                db.session.commit()
                flash("Nota criada.", "success")
                return redirect(url_for("notas_equipa"))
        notes = (
            EquipaNota.query.options(joinedload(EquipaNota.autor))
            .order_by(EquipaNota.updated_at.desc(), EquipaNota.created_at.desc())
            .all()
        )
        return render_template("notas_equipa.html", notes=notes)

    @app.route("/notas-equipa/<int:note_id>/editar", methods=["POST"])
    @login_required
    def atualizar_nota_equipa(note_id):
        note = EquipaNota.query.get_or_404(note_id)
        title = clean_text(request.form.get("titulo"))
        content = clean_text(request.form.get("conteudo"))
        if not title or not content:
            flash("Titulo e conteudo sao obrigatorios.", "error")
            return redirect(url_for("notas_equipa"))
        note.titulo = title
        note.conteudo = content
        note.updated_at = datetime.utcnow()
        db.session.commit()
        flash("Nota atualizada.", "success")
        return redirect(url_for("notas_equipa"))

    @app.route("/notas-equipa/<int:note_id>/apagar", methods=["POST"])
    @login_required
    def apagar_nota_equipa(note_id):
        note = EquipaNota.query.get_or_404(note_id)
        db.session.delete(note)
        db.session.commit()
        flash("Nota apagada.", "success")
        return redirect(url_for("notas_equipa"))

    @app.route("/admin")
    @login_required
    @admin_required
    def admin_overview():
        return render_template("admin_overview.html", **admin_overview_context())

    @app.route("/admin/users")
    @login_required
    @admin_required
    def admin_users():
        users = User.query.order_by(
            case((User.approval_status == "pending", 0), else_=1),
            User.ativo.desc(),
            User.role.asc(),
            User.nome.asc(),
        ).all()
        return render_template("admin_users.html", users=users)

    @app.route("/admin/users/<int:user_id>/approval", methods=["POST"])
    @login_required
    @admin_required
    def admin_user_approval(user_id):
        user = User.query.get_or_404(user_id)
        action = request.form.get("action")
        if user.id == current_user.id and action in {"reject", "deactivate"}:
            flash("Nao podes desativar o teu proprio utilizador.", "error")
            return redirect(url_for("admin_users"))
        if action == "approve":
            user.approval_status = "approved"
            user.ativo = True
            user.approved_at = datetime.utcnow()
            user.approved_by_id = current_user.id
            flash("Utilizador aprovado.", "success")
        elif action == "reject":
            user.approval_status = "rejected"
            user.ativo = False
            flash("Pedido de utilizador rejeitado.", "success")
        elif action == "deactivate":
            user.ativo = False
            flash("Utilizador desativado.", "success")
        elif action == "activate":
            if user.approval_status != "approved":
                flash("So podes ativar utilizadores aprovados.", "error")
                return redirect(url_for("admin_users"))
            user.ativo = True
            flash("Utilizador ativado.", "success")
        else:
            flash("Acao invalida.", "error")
            return redirect(url_for("admin_users"))
        db.session.commit()
        return redirect(url_for("admin_users"))

    @app.route("/meu-desempenho")
    @login_required
    def meu_desempenho():
        return render_template("meu_desempenho.html", performance=personal_performance_context(current_user))

    @app.route("/")
    @login_required
    def dashboard():
        if current_user.role == "admin":
            return redirect(url_for("admin_overview"))
        return redirect(url_for("mapa_leads"))

    @app.route("/mapa")
    @login_required
    def mapa_leads():
        start = time.perf_counter()
        options = get_options()
        metrics = minimal_map_metrics()
        safe_perf_log(f"[PERF] /mapa route {time.perf_counter() - start:.3f}s")
        return render_template(
            "mapa.html",
            options=options,
            metrics=metrics,
            assignment_scope=requested_assignment_scope(),
            assignment_scopes=assignment_scope_options(),
        )

    @app.route("/todas-leads")
    @login_required
    def todas_leads():
        scope = requested_assignment_scope()
        leads, payload, filters = all_leads_payload(scope)
        total = len(leads)
        with_coords = sum(1 for lead in leads if valid_coordinates(lead.latitude, lead.longitude))
        scheduled = sum(1 for lead in leads if is_scheduled_lead(lead))
        return render_template(
            "todas_leads.html",
            leads=payload,
            summary={
                "total": total,
                "with_coords": with_coords,
                "without_coords": total - with_coords,
                "scheduled": scheduled,
            },
            filters=filters,
            assignment_scope=scope,
            assignment_scopes=assignment_scope_options(),
        )

    @app.route("/todas-leads/exportar")
    @login_required
    def exportar_todas_leads():
        leads = all_leads_filtered_for_export(requested_assignment_scope())
        output = build_leads_export_workbook(leads)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"leads_alltera_{date.today().isoformat()}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/agendadas", methods=["GET", "POST"])
    @login_required
    def agendadas():
        if request.method == "POST":
            lead = Lead.query.get_or_404(int(request.form.get("lead_id")))
            action = request.form.get("action")
            note = clean_text(request.form.get("observacao"))
            if action == "contactada":
                lead.estado = "Ligar de volta"
                lead.estado_lead = lead.estado
                lead.data_novo_contacto = None
                lead.hora_reuniao = None
                add_history(lead, "Lead contactada", note or "Follow-up marcado como contactado.")
                flash("Lead marcada como contactada.", "success")
            elif action == "reagendar":
                new_date = parse_date(request.form.get("data_novo_contacto"))
                new_time = parse_time_value(request.form.get("hora_reuniao"))
                if not new_date:
                    flash("Escolhe uma data valida para reagendar.", "error")
                    return redirect(url_for("agendadas"))
                lead.estado = "Adiar contacto"
                lead.estado_lead = lead.estado
                lead.data_novo_contacto = new_date
                lead.hora_reuniao = new_time
                lead.motivo_classificacao = note or "Lead reagendada manualmente."
                time_note = f" às {new_time}" if new_time else ""
                add_history(lead, "Lead reagendada", f"Novo contacto: {new_date}{time_note}. {note}")
                flash("Lead reagendada.", "success")
            db.session.commit()
            return redirect(url_for("agendadas"))

        return render_template(
            "agendadas.html",
            scheduled=scheduled_leads_context(),
            today=date.today(),
            scheduled_datetime_label=scheduled_datetime_label,
            scheduled_time_label=scheduled_time_label,
        )

    @app.route("/leads/nova", methods=["GET", "POST"])
    @login_required
    @role_required("admin", "comercial", "rececionista")
    def adicionar_lead():
        edit_lead_id = request.form.get("lead_id", type=int) if request.method == "POST" else request.args.get("lead_id", type=int)
        edit_lead = Lead.query.get_or_404(edit_lead_id) if edit_lead_id else None
        template_options = {
            "lead_states": LEAD_STATES,
            "priority_options": ["Baixa", "Média", "Alta"],
            "tag_options": TAG_OPTIONS,
        }
        if request.method == "POST":
            lead_data = build_manual_lead(request.form)
            if not clean_text(lead_data["nome_cliente"] or lead_data["nome_empresa"]):
                flash("Indica o Nome Cliente ou a Empresa.", "error")
                return render_template("adicionar_lead.html", form=request.form, edit_lead=edit_lead, **template_options)
            if not clean_text(request.form.get("cidade") or request.form.get("morada") or request.form.get("codigo_postal")):
                flash("Indica pelo menos Cidade, Morada ou Código Postal.", "error")
                return render_template("adicionar_lead.html", form=request.form, edit_lead=edit_lead, **template_options)

            phone_index, fallback_index = build_duplicate_indexes(exclude_lead_id=edit_lead.id if edit_lead else None)
            duplicate, reason = detect_duplicate(
                lead_data,
                phone_index,
                fallback_index,
                exclude_lead_id=edit_lead.id if edit_lead else None,
            )
            if duplicate:
                action = "guardada" if edit_lead else "criada uma nova lead"
                flash(f"Lead duplicada encontrada ({reason}). Não foi {action}.", "error")
                return render_template("adicionar_lead.html", form=request.form, duplicate=duplicate, edit_lead=edit_lead, **template_options)

            if edit_lead:
                update_manual_lead(edit_lead, lead_data)
                add_history(edit_lead, "Lead editada manualmente", "Dados atualizados diretamente na aplicação.")
                db.session.commit()
                flash("Lead guardada com sucesso.", "success")
                return redirect(request.form.get("next") or url_for("mapa_leads"))

            lead = Lead(**lead_data)
            db.session.add(lead)
            db.session.flush()
            user_name = current_user.nome if current_user.is_authenticated else "utilizador"
            add_history(
                lead,
                "Lead criada manualmente",
                f"Lead criada manualmente por {user_name}",
                tipo_acao="Lead criada manualmente",
                user_id=current_user.id if current_user.is_authenticated else None,
            )
            db.session.commit()
            try:
                geocoded, message = geocode_single_lead(lead)
                if not geocoded:
                    add_history(lead, "Geocoding manual pendente", message)
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                lead = db.session.get(Lead, lead.id)
                add_history(lead, "Geocoding manual pendente", f"Não foi possível obter coordenadas: {exc}")
                db.session.commit()
            flash("Lead criada com sucesso.", "success")
            return redirect(request.form.get("next") or url_for("todas_leads"))

        prefill = {key: value for key, value in request.args.items() if value is not None}
        return render_template("adicionar_lead.html", form=prefill, edit_lead=edit_lead, **template_options)

    @app.route("/api/import/sheets", methods=["POST"])
    @login_required
    def api_import_sheets():
        file = request.files.get("ficheiro")
        if not file or not file.filename:
            return jsonify({"error": "Seleciona um ficheiro .xlsx."}), 400
        if not file.filename.lower().endswith(".xlsx"):
            return jsonify({"error": "A seleção de folhas só se aplica a ficheiros .xlsx."}), 400
        file_bytes = file.read()
        try:
            sheets = list_import_sheets(file_bytes)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400
        upload_id = uuid.uuid4().hex
        with IMPORT_UPLOADS_LOCK:
            IMPORT_UPLOADS[upload_id] = {
                "filename": file.filename,
                "bytes": file_bytes,
                "created_at": time.time(),
            }
        return jsonify({"upload_id": upload_id, "filename": file.filename, "sheets": sheets})

    @app.route("/api/import/start", methods=["POST"])
    @login_required
    def api_import_start():
        upload_id = request.form.get("upload_id")
        sheet_name = request.form.get("sheet_name") or None
        if upload_id:
            with IMPORT_UPLOADS_LOCK:
                upload = IMPORT_UPLOADS.pop(upload_id, None)
            if not upload:
                return jsonify({"error": "Upload expirado. Seleciona o ficheiro novamente."}), 400
            if not sheet_name:
                return jsonify({"error": "Seleciona uma folha antes de importar."}), 400
            filename = upload["filename"]
            file_bytes = upload["bytes"]
        else:
            file = request.files.get("ficheiro")
            if not file or not file.filename:
                return jsonify({"error": "Seleciona um ficheiro .xlsx ou .csv."}), 400
            filename = file.filename
            file_bytes = file.read()
            if filename.lower().endswith(".xlsx") and not sheet_name:
                return jsonify({"error": "Seleciona uma folha antes de importar."}), 400
        auto_geocode = request.form.get("auto_geocode") == "1"
        job_id = create_import_job(user_id=current_user.id)
        thread = threading.Thread(
            target=run_import_job,
            args=(app, job_id, filename, file_bytes, auto_geocode, sheet_name, current_user.id),
            daemon=True,
        )
        thread.start()
        return jsonify({"job_id": job_id})

    @app.route("/api/import/progress/<job_id>")
    @login_required
    def api_import_progress(job_id):
        job = get_import_job(job_id)
        if not job:
            return jsonify({"error": "Importacao nao encontrada."}), 404
        return jsonify(job)

    @app.route("/importar", methods=["GET", "POST"])
    @login_required
    def importar_leads():
        summary = None
        mapping_preview = None
        if request.method == "POST":
            print("[IMPORT] INICIO IMPORTACAO", flush=True)
            try:
                file = request.files.get("ficheiro")
                if not file or not file.filename:
                    print("[IMPORT] erro: ficheiro nao enviado", flush=True)
                    flash("Seleciona um ficheiro .xlsx ou .csv.", "error")
                    return redirect(url_for("importar_leads"))

                print(f"[IMPORT] ficheiro recebido: {file.filename}", flush=True)
                rows = read_import_file(file, sheet_name=request.form.get("sheet_name") or None)
                print(f"[IMPORT] linhas lidas: {len(rows)}", flush=True)
                headers = list(rows[0].keys()) if rows else []
                mapping = map_columns(headers)
                print(f"[IMPORT] colunas detetadas: {mapping}", flush=True)
                parsed_rows = parse_rows(rows, mapping)
                auto_geocode = request.form.get("auto_geocode", "1") == "1"
                summary = import_summary(parsed_rows, auto_geocode=auto_geocode)
                if request.form.get("sheet_name"):
                    summary["sheets_imported"] = [request.form.get("sheet_name")]
                print(
                    f"[IMPORT] FIM IMPORTACAO criadas={summary['imported']} "
                    f"atualizadas={summary['updated']} duplicados={summary['duplicates']} erros={len(summary['errors'])}",
                    flush=True,
                )
                flash("Importacao concluida.", "success")
            except OperationalError as exc:
                db.session.rollback()
                print(f"[IMPORT] erro operacional: {exc}", flush=True)
                flash(str(exc), "error")
            except Exception as exc:
                db.session.rollback()
                print(f"[IMPORT] erro: {exc}", flush=True)
                flash(str(exc), "error")
            finally:
                db.session.remove()
                print("[IMPORT] sessao fechada", flush=True)
        return render_template("importar.html", columns=IMPORT_COLUMNS, real_columns=ALLTERA_REAL_COLUMNS, summary=summary, mapping_preview=mapping_preview, latest_import=latest_import_summary())

    @app.route("/planeamento", methods=["GET", "POST"])
    @login_required
    def planeamento():
        day_plan = None
        plan_summary = None
        displacement = []
        meetings = []
        radius = float(request.form.get("raio", 10) if request.method == "POST" else request.args.get("raio", 10))
        selected_commercial = request.form.get("comercial", PLANNING_COMMERCIALS[0]) if request.method == "POST" else request.args.get("comercial", PLANNING_COMMERCIALS[0])
        if selected_commercial not in PLANNING_COMMERCIALS:
            selected_commercial = PLANNING_COMMERCIALS[0]
        selected_city = request.form.get("cidade", "") if request.method == "POST" else request.args.get("cidade", "")
        selected_type = request.form.get("tipo", "") if request.method == "POST" else request.args.get("tipo", "")
        selected_base_lead_id = int(request.form.get("lead_base") or 0) if request.method == "POST" else int(request.args.get("lead_base") or 0)
        limit = int(request.form.get("limite", 12) if request.method == "POST" else request.args.get("limite", 12))

        eligible_filtered = [
            lead for lead in Lead.query.all()
            if eligible_planning_lead(lead)
            and (not selected_city or (lead.cidade or lead.localidade) == selected_city)
            and (not selected_type or (lead.area_negocio or lead.tipo_cliente) == selected_type)
        ]
        planning_base_leads = [
            lead for lead in eligible_filtered
            if commercial_matches(lead, selected_commercial)
        ]

        if request.method == "POST":
            candidates = planning_base_leads
            excluded_other_count = len(eligible_filtered) - len(candidates)
            day_plan = generate_day_contact_plan(
                candidates,
                radius_km=radius,
                limit=limit,
                base_lead_id=selected_base_lead_id or None,
                base_city=selected_city,
            )
            day_plan["commercial"] = selected_commercial
            plan_summary = day_planning_summary(candidates, day_plan, excluded_other_count=excluded_other_count)
        return render_template(
            "planeamento.html",
            options=get_options(),
            meetings=meetings,
            displacement=displacement,
            day_plan=day_plan,
            plan_summary=plan_summary,
            planning_base_leads=planning_base_leads,
            planning_comerciais=PLANNING_COMMERCIALS,
            unassigned_commercial=UNASSIGNED_COMMERCIAL,
            planning_ownership_label=planning_ownership_label,
            filters={
                "comercial": selected_commercial,
                "cidade": selected_city,
                "tipo": selected_type,
                "raio": radius,
                "lead_base": selected_base_lead_id,
                "limite": limit,
            },
        )

    @app.route("/agenda")
    @login_required
    def agenda():
        flash("Agenda removida: usa o Plano do dia para contactos geográficos.", "info")
        return redirect(url_for("planeamento"))

    @app.route("/lista-dia")
    @login_required
    def lista_dia():
        flash("Lista de dia antiga removida: usa o Plano do dia operacional.", "info")
        return redirect(url_for("planeamento"))

    @app.route("/sem-coordenadas", methods=["GET", "POST"])
    @login_required
    def sem_coordenadas():
        if request.method == "POST":
            action = request.form.get("action")
            if action == "geocode_all":
                success = 0
                failed = 0
                for lead in Lead.query.all():
                    if needs_coordinate_review(lead) and "Ignorar mapa" not in lead.tag_list():
                        ok, _ = geocode_single_lead(lead)
                        success += 1 if ok else 0
                        failed += 0 if ok else 1
                db.session.commit()
                flash(f"Geocoding concluido: {success} resolvidas, {failed} por rever.", "success" if failed == 0 else "warning")
                return redirect(url_for("sem_coordenadas"))

            selected_ids = [int(value) for value in request.form.getlist("lead_ids") if str(value).isdigit()]
            if action in {"bulk_retry", "bulk_ignore", "bulk_resolved"}:
                selected = Lead.query.filter(Lead.id.in_(selected_ids)).all() if selected_ids else []
                success = 0
                for lead in selected:
                    if action == "bulk_retry":
                        ok, _ = geocode_single_lead(lead)
                        success += 1 if ok else 0
                    elif action == "bulk_ignore":
                        tags = set(lead.tag_list())
                        tags.add("Ignorar mapa")
                        lead.tags = ", ".join(sorted(tags))
                        lead.latitude = None
                        lead.longitude = None
                        add_history(lead, "Ignorar no mapa", "Lead marcada em lote para ficar fora do mapa.")
                        success += 1
                    elif action == "bulk_resolved" and valid_coordinates(lead.latitude, lead.longitude):
                        tags = [tag for tag in lead.tag_list() if tag != "Ignorar mapa"]
                        lead.tags = ", ".join(tags)
                        add_history(lead, "Coordenadas resolvidas", "Lead marcada em lote como resolvida.")
                        success += 1
                db.session.commit()
                flash(f"Ação em lote aplicada a {success} leads.", "success")
                return redirect(url_for("sem_coordenadas"))

            lead = Lead.query.get_or_404(int(request.form.get("lead_id")))
            if action == "update_location":
                lead.cidade = normalize_city(request.form.get("cidade")) or "Sem cidade"
                lead.localidade = lead.cidade
                lead.morada = clean_text(request.form.get("morada"))
                lead.codigo_postal = clean_text(request.form.get("codigo_postal"))
                lead.latitude = None
                lead.longitude = None
                add_history(lead, "Localizacao corrigida", f"Cidade: {lead.cidade}. Morada: {lead.morada}. CP: {lead.codigo_postal}.")
                flash("Localizacao atualizada.", "success")
            elif action == "retry_geocode":
                ok, message = geocode_single_lead(lead)
                flash(message, "success" if ok else "warning")
            elif action == "ignore_map":
                tags = set(lead.tag_list())
                tags.add("Ignorar mapa")
                lead.tags = ", ".join(sorted(tags))
                lead.latitude = None
                lead.longitude = None
                add_history(lead, "Ignorar no mapa", "Lead marcada para revisao operacional fora do mapa.")
                flash("Lead marcada para ignorar no mapa.", "info")
            db.session.commit()
            return redirect(url_for("sem_coordenadas"))

        status_filter = request.args.get("estado", "todas")
        leads = [lead for lead in Lead.query.order_by(Lead.nome_empresa.asc()).all() if needs_coordinate_review(lead)]
        if status_filter == "ativas":
            leads = [lead for lead in leads if is_active_lead(lead)]
        elif status_filter == "inativas":
            leads = [lead for lead in leads if is_inactive_lead(lead)]
        rows = [{"lead": lead, "motivo": coordinate_review_reason(lead), "sugestao": city_suggestion(lead.cidade or lead.localidade)} for lead in leads]
        return render_template("sem_coordenadas.html", rows=rows, status_filter=status_filter, quality=coordinate_quality_context(rows))

    @app.route("/leads-inativas", methods=["GET", "POST"])
    @login_required
    def leads_inativas():
        if request.method == "POST":
            lead = Lead.query.get_or_404(int(request.form.get("lead_id")))
            action = request.form.get("action")
            if action == "reativar":
                lead.estado = "Por contactar"
                lead.estado_lead = "Por contactar"
                lead.data_novo_contacto = None
                lead.hora_reuniao = None
                lead.classificacao_observacao = "Corrigido manualmente"
                lead.motivo_classificacao = "Lead reativada manualmente após revisão."
                add_history(lead, "Lead reativada", "Estado alterado para Por contactar.")
                flash("Lead reativada.", "success")
            elif action == "corrigir_estado":
                new_state = clean_text(request.form.get("estado"))
                if new_state in LEAD_STATES:
                    lead.estado = new_state
                    lead.estado_lead = new_state
                    lead.classificacao_observacao = "Corrigido manualmente"
                    lead.motivo_classificacao = f"Estado corrigido manualmente para {new_state}."
                    if new_state != "Adiar contacto":
                        lead.data_novo_contacto = None
                        lead.hora_reuniao = None
                    add_history(lead, "Estado corrigido manualmente", lead.motivo_classificacao)
                    flash("Estado atualizado.", "success")
                else:
                    flash("Estado inválido.", "error")
            elif action == "adiar":
                lead.estado = "Adiar contacto"
                lead.estado_lead = "Adiar contacto"
                lead.data_novo_contacto = parse_date(request.form.get("data_novo_contacto"))
                lead.hora_reuniao = parse_time_value(request.form.get("hora_reuniao"))
                lead.motivo_classificacao = "Contacto adiado manualmente na revisão de inativas."
                time_note = f" às {lead.hora_reuniao}" if lead.hora_reuniao else ""
                add_history(lead, "Adiar contacto", f"Novo contacto: {lead.data_novo_contacto or ''}{time_note}.")
                flash("Contacto adiado.", "success")
            db.session.commit()
            return redirect(url_for("leads_inativas"))

        filters = {
            "estado": request.args.get("estado", ""),
            "cidade": request.args.get("cidade", ""),
            "comercial": request.args.get("comercial", ""),
            "texto": request.args.get("texto", ""),
            "motivo": request.args.get("motivo", ""),
        }
        leads = [lead for lead in Lead.query.order_by(Lead.nome_empresa.asc()).all() if is_inactive_lead(lead)]
        if filters["estado"]:
            leads = [lead for lead in leads if lead.estado == filters["estado"]]
        if filters["cidade"]:
            leads = [lead for lead in leads if normalize_lookup(lead.cidade or lead.localidade) == normalize_lookup(filters["cidade"])]
        if filters["comercial"]:
            leads = [lead for lead in leads if lead.comercial_responsavel == filters["comercial"]]
        if filters["motivo"]:
            leads = [lead for lead in leads if filters["motivo"] in {lead.classificacao_observacao, lead.motivo_classificacao}]
        if filters["texto"]:
            needle = normalize_lookup(filters["texto"])
            leads = [
                lead for lead in leads
                if needle in normalize_lookup(" ".join([
                    lead.nome_cliente or "",
                    lead.nome_empresa or "",
                    lead.telefone or "",
                    lead.email or "",
                    lead.observacoes or "",
                    lead.observacoes_contacto or "",
                ]))
            ]

        options = {
            "estados": sorted({lead.estado for lead in Lead.query.all() if lead.estado}),
            "cidades": sorted({lead.cidade or lead.localidade for lead in Lead.query.all() if lead.cidade or lead.localidade}),
            "comerciais": sorted({lead.comercial_responsavel for lead in Lead.query.all() if lead.comercial_responsavel}),
            "motivos": sorted({lead.classificacao_observacao for lead in Lead.query.all() if lead.classificacao_observacao}),
        }
        return render_template("leads_inativas.html", leads=leads, filters=filters, options=options, estados=LEAD_STATES)

    @app.route("/esquecidas")
    @login_required
    def esquecidas():
        days = int(request.args.get("dias", 60) or 60)
        commercial = request.args.get("comercial", "")
        city = request.args.get("cidade", "")
        query = normalize_lookup(request.args.get("q", ""))
        rows = forgotten_leads(days)
        if commercial:
            rows = [(lead, latest, age) for lead, latest, age in rows if normalize_commercial_key(lead.comercial_responsavel) == commercial]
        if city:
            rows = [(lead, latest, age) for lead, latest, age in rows if (lead.cidade or lead.localidade) == city]
        if query:
            rows = [
                (lead, latest, age)
                for lead, latest, age in rows
                if query in normalize_lookup(" ".join([lead.nome_cliente or "", lead.nome_empresa or "", lead.telefone or "", lead.email or ""]))
            ]
        return render_template(
            "esquecidas.html",
            options=get_options(),
            forgotten=rows,
            filters={"dias": days, "comercial": commercial, "cidade": city, "q": request.args.get("q", "")},
        )

    @app.route("/duplicados", methods=["GET", "POST"])
    @login_required
    def duplicados():
        if request.method == "POST":
            duplicate = PossivelDuplicado.query.get_or_404(int(request.form["id"]))
            action = request.form["action"]
            imported = duplicate.imported_data()
            if action == "fundir":
                if update_empty_fields(duplicate.lead, imported):
                    add_history(duplicate.lead, "Duplicado fundido", f"Dados aproximados integrados: {imported.get('nome_empresa')}")
                duplicate.estado = "Fundido"
            elif action == "manter":
                imported.setdefault("comercial_responsavel", "Outro")
                lead = Lead(**imported, estado="Por contactar")
                db.session.add(lead)
                duplicate.estado = "Mantido separado"
            else:
                duplicate.estado = "Ignorado"
            db.session.commit()
            return redirect(url_for("duplicados"))
        candidates = PossivelDuplicado.query.order_by(PossivelDuplicado.created_at.desc()).all()
        return render_template("duplicados.html", candidates=candidates)

    @app.route("/historico")
    @login_required
    def historico():
        lead_id = request.args.get("lead_id", type=int)
        query = HistoricoLead.query
        if lead_id:
            query = query.filter_by(lead_id=lead_id)
        entries = (
            query.options(joinedload(HistoricoLead.user), joinedload(HistoricoLead.lead))
            .order_by(HistoricoLead.created_at.desc())
            .limit(300)
            .all()
        )
        return render_template("historico.html", entries=entries)

    @app.route("/api/leads")
    @login_required
    def api_leads():
        start = time.perf_counter()
        include_history = request.args.get("history") == "1"
        lite = request.args.get("lite") == "1"
        scope = requested_assignment_scope()
        bounds = {
            key: parse_float_optional(request.args.get(key))
            for key in ("north", "south", "east", "west")
        }
        has_bounds = all(value is not None for value in bounds.values())
        selected_lead_id = request.args.get("lead_id", type=int)
        query = apply_assignment_scope(Lead.query, scope)
        if lite:
            query = aplicar_filtro_leads_operacionais_mapa(query)
        if not lite:
            query = query.options(joinedload(Lead.assigned_to), selectinload(Lead.historico))
        elif include_history:
            query = query.options(selectinload(Lead.historico))
        if has_bounds:
            spatial_filters = (
                Lead.latitude.isnot(None),
                Lead.longitude.isnot(None),
                Lead.latitude <= bounds["north"],
                Lead.latitude >= bounds["south"],
            )
            if bounds["east"] >= bounds["west"]:
                longitude_filter = Lead.longitude.between(bounds["west"], bounds["east"])
            else:
                longitude_filter = or_(Lead.longitude >= bounds["west"], Lead.longitude <= bounds["east"])
            spatial_filter = and_(*spatial_filters, longitude_filter)
            if selected_lead_id:
                query = query.filter(or_(spatial_filter, Lead.id == selected_lead_id))
            else:
                query = query.filter(spatial_filter)
        leads = query.order_by(Lead.nome_empresa.asc()).all()
        query_elapsed = time.perf_counter() - start
        payload = []
        for lead in leads:
            if lite:
                item = lead_map_payload(lead, include_history=include_history)
            else:
                item = lead.to_dict(include_history=include_history)
                item["estado"] = normalize_legacy_state(item["estado"])
                item["comercial_responsavel"] = display_commercial(item.get("comercial_responsavel"))
                item["comercial_key"] = normalize_commercial_key(item.get("comercial_responsavel"))
                item["ativa"] = is_active_lead(lead)
                item["agendada"] = is_scheduled_lead(lead)
                item["agenda_bucket"] = scheduled_bucket(lead)
                item["tem_coordenadas"] = lead.latitude is not None and lead.longitude is not None
            payload.append(item)
        safe_perf_log(
            f"[PERF] /api/leads count={len(leads)} bounds={has_bounds} lite={lite} query={query_elapsed:.3f}s total={time.perf_counter() - start:.3f}s"
        )
        return jsonify(payload)

    @app.route("/api/leads/<int:lead_id>/resumo")
    @login_required
    def api_lead_resumo(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        item = lead_map_payload(lead, include_history=False)
        item.update({
            "nome_cliente": lead.nome_cliente or lead.nome_empresa or "",
            "nome_empresa": lead.nome_empresa or "",
            "empresa": lead.empresa or "",
            "area_negocio": lead.area_negocio or lead.tipo_cliente or "",
            "tipo_cliente": lead.tipo_cliente or "",
            "telefone": lead.telefone or lead.contacto or "",
            "email": lead.email or "",
            "cidade": lead.cidade or lead.localidade or "",
            "localidade": lead.localidade or "",
            "morada": lead.morada or "",
            "codigo_postal": lead.codigo_postal or "",
            "contacto": lead.contacto or "",
            "categoria": lead.categoria or "",
            "nif": lead.nif or "",
            "observacoes": lead.observacoes or "",
            "observacoes_contacto": lead.observacoes_contacto or "",
            "reuniao_info": lead.reuniao_info or "",
            "classificacao_observacao": lead.classificacao_observacao or "",
            "motivo_classificacao": lead.motivo_classificacao or "",
            "insight_note": lead.insight_note or "",
        })
        # Resumo leve para o drawer: a timeline completa continua fora do payload inicial do mapa.
        history = (
            HistoricoLead.query
            .options(joinedload(HistoricoLead.user))
            .filter(HistoricoLead.lead_id == lead.id)
            .order_by(HistoricoLead.created_at.desc())
            .limit(5)
            .all()
        )
        timeline = []
        for entry in history:
            timeline.append({
                "id": entry.id,
                "lead_id": entry.lead_id,
                "titulo": entry.acao or "Evento",
                "acao": entry.acao or "Evento",
                "tipo_acao": entry.tipo_acao or "",
                "observacao": entry.observacao or "",
                "descricao": entry.observacao or entry.resultado or "",
                "resultado": entry.resultado or "",
                "utilizador": entry.user.nome if entry.user else "",
                "user": entry.user.nome if entry.user else "",
                "created_at": entry.created_at.strftime("%d/%m/%Y %H:%M") if entry.created_at else "",
            })
        item["timeline"] = timeline
        item["historico"] = timeline
        return jsonify(item)

    @app.route("/api/search")
    @login_required
    def api_search():
        return jsonify({"results": search_leads(request.args.get("q", ""), limit=10)})

    @app.route("/api/leads/next")
    @login_required
    def api_next_lead():
        base_id = request.args.get("base_id", type=int)
        commercial = request.args.get("commercial")
        base_lead = db.session.get(Lead, base_id) if base_id else None
        if base_id and not base_lead:
            return jsonify({"error": "Lead base nao encontrada."}), 404
        lead = next_best_lead(base_lead=base_lead, commercial_filter=commercial)
        return jsonify({"lead": lead.to_dict(include_history=True) if lead else None})

    @app.route("/api/leads/bulk-action", methods=["POST"])
    @login_required
    def api_leads_bulk_action():
        data = request.get_json(silent=True) or {}
        ids = [int(value) for value in data.get("ids", []) if str(value).isdigit()]
        action = data.get("action")
        query = Lead.query.filter(Lead.id.in_(ids)) if ids else Lead.query.filter(db.text("0=1"))
        leads = query.all()
        if not leads:
            return jsonify({"updated": 0})
        for lead in leads:
            if action == "assign_commercial":
                lead.comercial_responsavel = commercial_label_from_key(data.get("comercial"))
                add_history(lead, "Comercial atribuído", f"Atribuído em lote a {lead.comercial_responsavel}.")
            elif action == "assign_user":
                if current_user.role != "admin":
                    return jsonify({"error": "Sem permissao para atribuir leads"}), 403
                assigned_to_id = data.get("assigned_to_id")
                user = db.session.get(User, int(assigned_to_id)) if str(assigned_to_id or "").isdigit() else None
                if assigned_to_id and not user:
                    return jsonify({"error": "Utilizador invalido"}), 400
                lead.assigned_to = user
                add_history(
                    lead,
                    "Lead atribuida",
                    f"Atribuida a {user.nome}." if user else "Atribuicao removida.",
                    tipo_acao="edicao_lead",
                )
            elif action == "clear_commercial":
                lead.comercial_responsavel = UNASSIGNED_COMMERCIAL
                add_history(lead, "Comercial removido", "Comercial removido em lote.")
            elif action == "set_state":
                state = clean_text(data.get("estado"))
                if state not in LEAD_STATES:
                    return jsonify({"error": "Estado inválido"}), 400
                lead.estado = state
                lead.estado_lead = state
                add_history(lead, "Estado alterado", f"Estado alterado em lote para {state}.")
            elif action == "ignore_map":
                tags = set(lead.tag_list())
                tags.add("Ignorar mapa")
                lead.tags = ", ".join(sorted(tags))
                lead.latitude = None
                lead.longitude = None
                add_history(lead, "Ignorar no mapa", "Lead marcada em lote para ficar fora do mapa.")
            else:
                return jsonify({"error": "Ação inválida"}), 400
        db.session.commit()
        return jsonify({"updated": len(leads)})

    @app.route("/api/leads/<int:lead_id>/action", methods=["POST"])
    @login_required
    def lead_action(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        data = request.get_json(silent=True) or {}
        action = data.get("action")
        commercial = data.get("comercial_responsavel") or lead.comercial_responsavel
        observation = clean_text(data.get("observacao"))

        if action in {"contactado", "ligar_volta"}:
            lead.estado = "Ligar de volta"
            lead.estado_lead = "Ligar de volta"
            if action == "contactado":
                lead.data_novo_contacto = None
                lead.hora_reuniao = None
            else:
                lead.data_novo_contacto = parse_date(data.get("data_novo_contacto"))
                lead.hora_reuniao = parse_time_value(data.get("hora_reuniao"))
            add_history(lead, "Estado alterado", observation or "Lead marcada para ligar de volta.", commercial)
        elif action == "corrigir_estado":
            new_state = clean_text(data.get("estado"))
            if new_state not in LEAD_STATES:
                return jsonify({"error": "Estado invalido"}), 400
            lead.estado = new_state
            lead.classificacao_observacao = "Corrigido manualmente"
            lead.motivo_classificacao = observation or f"Estado corrigido manualmente para {new_state}."
            if new_state != "Adiar contacto":
                lead.data_novo_contacto = None
                lead.hora_reuniao = None
            add_history(lead, "Estado corrigido manualmente", lead.motivo_classificacao, commercial)
        elif action in {"reuniao", "cliente_existente", "crm"}:
            lead.estado = "Já tratado / no CRM"
            lead.data_reuniao = None
            lead.hora_reuniao = None
            add_history(lead, "Reunião/CRM registado", observation or "Lead já tratada no CRM.", commercial, tipo_acao="reuniao_marcada")
        elif action == "adiar":
            new_contact_date = parse_date(data.get("data_novo_contacto"))
            new_contact_time = parse_time_value(data.get("hora_reuniao"))
            if not new_contact_date:
                return jsonify({"error": "Data de novo contacto obrigatoria"}), 400
            lead.estado = "Adiar contacto"
            lead.estado_lead = "Adiar contacto"
            lead.comercial_responsavel = commercial
            lead.data_novo_contacto = new_contact_date
            lead.hora_reuniao = new_contact_time
            time_note = f" às {new_contact_time}" if new_contact_time else ""
            add_history(lead, "Adiar contacto", f"Novo contacto: {lead.data_novo_contacto or ''}{time_note}. {observation}", commercial, tipo_acao="followup_reagendado")
        elif action == "sem_interesse":
            new_contact_date = parse_date(data.get("data_novo_contacto"))
            new_contact_time = parse_time_value(data.get("hora_reuniao"))
            temporary_interest_tokens = [
                "não agora",
                "nao agora",
                "sem interesse momentâneo",
                "sem interesse momentaneo",
                "contactar mais tarde",
                "ligar mais tarde",
                "voltar a ligar",
                "voltar a contactar",
            ]
            observation_lookup = normalize_lookup(observation)
            if new_contact_date or any(token in observation_lookup for token in temporary_interest_tokens):
                if not new_contact_date:
                    return jsonify({"error": "Data de novo contacto obrigatoria"}), 400
                lead.estado = "Adiar contacto"
                lead.estado_lead = "Adiar contacto"
                lead.comercial_responsavel = commercial
                lead.data_novo_contacto = new_contact_date
                lead.hora_reuniao = new_contact_time
                time_note = f" às {new_contact_time}" if new_contact_time else ""
                add_history(lead, "Adiar contacto", f"Novo contacto: {lead.data_novo_contacto or ''}{time_note}. {observation}", commercial, tipo_acao="followup_reagendado")
            else:
                lead.estado = "Sem interesse"
                lead.estado_lead = "Sem interesse"
                lead.data_novo_contacto = None
                lead.hora_reuniao = None
                add_history(lead, "Estado alterado", observation or "Lead marcada sem interesse definitivo.", commercial)
        elif action == "atribuir":
            lead.comercial_responsavel = commercial
            add_history(lead, "Comercial atribuido", observation, commercial)
        elif action == "assign_user":
            if current_user.role != "admin":
                return jsonify({"error": "Sem permissao para atribuir leads"}), 403
            assigned_to_id = data.get("assigned_to_id")
            user = db.session.get(User, int(assigned_to_id)) if str(assigned_to_id or "").isdigit() else None
            if assigned_to_id and not user:
                return jsonify({"error": "Utilizador invalido"}), 400
            lead.assigned_to = user
            add_history(
                lead,
                "Lead atribuida",
                f"Atribuida a {user.nome}." if user else "Atribuicao removida.",
                tipo_acao="edicao_lead",
            )
        elif action == "add_note":
            if not observation:
                return jsonify({"error": "Nota vazia"}), 400
            add_history(lead, "Nota adicionada", observation, commercial)
        elif action == "day_plan":
            add_history(lead, "Adicionada a plano do dia", observation or "Lead incluida na rota do dia.", commercial)
        elif action == "add_tag":
            tag = clean_text(data.get("tag"))
            tags = set(lead.tag_list())
            if tag:
                tags.add(tag)
                lead.tags = ", ".join(sorted(tags))
                add_history(lead, "Tag adicionada", tag, commercial)
        elif action == "remove_tag":
            tag = clean_text(data.get("tag"))
            tags = [item for item in lead.tag_list() if item != tag]
            lead.tags = ", ".join(tags)
            add_history(lead, "Tag removida", tag, commercial)
        elif action == "update_insights":
            requested_tags = data.get("insight_tags") or []
            if not isinstance(requested_tags, list):
                requested_tags = []
            old_tags = set(lead.insight_tag_list())
            old_note = clean_text(lead.insight_note)
            allowed_tags = set(INSIGHT_TAG_OPTIONS)
            new_tags = [clean_text(tag) for tag in requested_tags if clean_text(tag) in allowed_tags]
            note = clean_text(data.get("insight_note"))
            lead.insight_tags = ", ".join(new_tags)
            lead.insight_note = note
            added = sorted(set(new_tags) - old_tags)
            removed = sorted(old_tags - set(new_tags))
            if added:
                add_history(lead, "Insight tags adicionadas", ", ".join(added), commercial)
            if removed:
                add_history(lead, "Insight tags removidas", ", ".join(removed), commercial)
            if note:
                action_label = "Insight interno criado" if not old_note and not old_tags else "Insight interno editado"
                add_history(lead, action_label, note, commercial)
            elif added or removed:
                add_history(lead, "Insight interno atualizado", "Tags internas atualizadas.", commercial)
        elif action == "update_coordinates":
            latitude = parse_float_optional(data.get("latitude"))
            longitude = parse_float_optional(data.get("longitude"))
            if latitude is None or longitude is None:
                return jsonify({"error": "Coordenadas invalidas"}), 400
            lead.latitude = latitude
            lead.longitude = longitude
            add_history(lead, "Coordenadas completadas", f"Latitude: {latitude}; Longitude: {longitude}", commercial)
        else:
            return jsonify({"error": "Acao invalida"}), 400

        lead.estado_lead = lead.estado
        db.session.commit()
        return jsonify({**lead.to_dict(include_history=True), "ativa": is_active_lead(lead)})

    @app.route("/api/export-plan", methods=["POST"])
    @login_required
    def export_plan():
        rows = (request.get_json(silent=True) or {}).get("rows", [])
        if rows:
            first = rows[0]
            db.session.add(PlanoReunioes(comercial=first.get("comercial", "Alltera"), data=parse_date(first.get("data")) or date.today(), total_leads=len(rows), observacao="Lista de contactos criada."))
            lead = Lead.query.filter_by(nome_empresa=first.get("nome_empresa")).first()
            if lead:
                add_history(lead, "Lista de contactos criada", f"{len(rows)} contactos exportados.", first.get("comercial"))
            db.session.commit()
        output = export_workbook(rows)
        return send_file(
            output,
            as_attachment=True,
            download_name="lista_contactos_zona_alltera.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


app = create_app()


if __name__ == "__main__":
    app.run(debug=True)
